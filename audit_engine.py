#!/usr/bin/env python3
"""
audit_engine.py — Motor genérico de auditoría SEO (Screaming Frog + GSC)

Uso:
  python3 audit_engine.py --config config_newcop.py
  python3 audit_engine.py --config config_clienteX.py --csv /otra/ruta.csv
  python3 audit_engine.py --config config_clienteX.py --output-dir /tmp/

Requiere: profiler_csv.py en el mismo directorio
"""

import sys
import re
import argparse
import importlib.util
import warnings
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore', category=UserWarning)


# ──────────────────────────────────────────────────────────────────────────────
# 1. CLI + CONFIG
# ──────────────────────────────────────────────────────────────────────────────

def load_config(config_path):
    spec = importlib.util.spec_from_file_location('cfg', config_path)
    cfg = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(cfg)
    return cfg


# ──────────────────────────────────────────────────────────────────────────────
# MOTOR PRINCIPAL — llamable desde CLI y desde la UI web
# ──────────────────────────────────────────────────────────────────────────────

def run_audit(cfg, ruta_csv, output_path):
    """
    Ejecuta la auditoría SEO completa.

    Args:
        cfg           -- módulo/objeto con los atributos de configuración del cliente
        ruta_csv      -- ruta absoluta al export CSV de Screaming Frog
        output_path   -- ruta completa del Excel de salida (.xlsx)

    Returns:
        dict con claves: tasks, urls, gsc, resumen, output_path
    """
    RUTA_CSV   = ruta_csv
    OUTPUT_PATH = output_path
    DOMAIN     = cfg.DOMAIN
    PLATFORM   = getattr(cfg, 'PLATFORM', 'Generic')

    SITE_LANGS      = getattr(cfg, 'SITE_LANGS', [])
    IS_MULTILINGUAL = getattr(cfg, 'IS_MULTILINGUAL', bool(SITE_LANGS))
    SPECIAL_LOCALE  = getattr(cfg, 'SPECIAL_LOCALE', None)

    URL_PATTERNS    = getattr(cfg, 'URL_PATTERNS', [])
    SYSTEM_PATTERNS = getattr(cfg, 'SYSTEM_PATTERNS', [])
    SEO_TYPES       = getattr(cfg, 'SEO_TYPES', ['product', 'blog', 'page'])
    HAS_CART_RECO   = getattr(cfg, 'HAS_CART_RECO', False)

    PAGINATION_PARAMS     = getattr(cfg, 'PAGINATION_PARAMS', ['page'])
    PAGINATION_PATH_REGEX = getattr(cfg, 'PAGINATION_PATH_REGEX', None)

    THRESHOLDS = getattr(cfg, 'THRESHOLDS', {})
    T_IMPRESSIONS_DEMAND   = THRESHOLDS.get('impressions_demand', 200)
    T_INLINKS_LOW          = THRESHOLDS.get('inlinks_low', 2)
    T_CTR_LOW              = THRESHOLDS.get('ctr_low', 0.01)
    T_IMPRESSIONS_MIN_CTR  = THRESHOLDS.get('impressions_min_ctr', 50)
    T_IMPRESSIONS_NO_CLICKS= THRESHOLDS.get('impressions_no_clicks', 500)
    T_POS_OPP_MIN          = THRESHOLDS.get('pos_opportunity_min', 11)
    T_POS_OPP_MAX          = THRESHOLDS.get('pos_opportunity_max', 20)
    T_TITLE_SHORT_CHARS    = THRESHOLDS.get('title_short_chars', 40)
    T_INLINKS_4XX_CRITICAL = THRESHOLDS.get('inlinks_4xx_critical', 500)
    T_WORD_COUNT_THIN      = THRESHOLDS.get('word_count_thin', 150)
    T_ORPHAN_IMPRESSIONS   = THRESHOLDS.get('orphan_impressions', 100)

    HEADER_BG = getattr(cfg, 'HEADER_BG', '1F4E78')
    P0_BG     = getattr(cfg, 'P0_BG', 'FFDCE0')
    P1_BG     = getattr(cfg, 'P1_BG', 'FFF0D3')
    P2_BG     = getattr(cfg, 'P2_BG', 'FFFACC')
    P3_BG     = getattr(cfg, 'P3_BG', 'E8F5E9')
    ALT_ROW   = getattr(cfg, 'ALT_ROW', 'F5F8FC')
    WHITE     = 'FFFFFF'

    print(f"=== AUDIT ENGINE ===")
    print(f"CSV     : {RUTA_CSV}")
    print(f"Output  : {OUTPUT_PATH}")
    print(f"Platform: {PLATFORM}")



    try:
        import profiler_csv   # noqa: E402  (same directory, dev only)
        print("\nProfiling CSV...")
        pf = profiler_csv.profile(RUTA_CSV, emit_json=False)
    except ModuleNotFoundError:
        pf = {'rename': {}, 'decimal_comma': False, 'has_gsc': False, 'lang': 'en', 'ctr_scale': '0-1'}

    RENAME       = pf['rename']
    DECIMAL_COMMA = pf['decimal_comma']
    HAS_GSC      = pf['has_gsc']
    SF_LANG      = pf['lang']

    print(f"\nLoading CSV ({RUTA_CSV.split('/')[-1]})...")
    df_raw = pd.read_csv(RUTA_CSV, low_memory=False, encoding='utf-8-sig')
    # Limpiar BOM y espacios en nombres de columna
    df_raw.columns = [c.strip().lstrip('\ufeff') for c in df_raw.columns]
    df_raw.rename(columns=RENAME, inplace=True)

    # Normalización de columnas Screaming Frog (fallback cuando profiler_csv no aplica)
    SF_COL_MAP = {
        'Address': 'url',
        'Status Code': 'status',
        'Indexability': 'indexable',
        'Title 1': 'title',
        'Title 1 Length': 'title_len',
        'Meta Description 1': 'meta_desc',
        'Meta Description 1 Length': 'meta_desc_len',
        'H1-1': 'h1',
        'H2-1': 'h2',
        'Canonical Link Element 1': 'canonical',
        'Meta Robots 1': 'meta_robots',
        'Crawl Depth': 'depth',
        'Inlinks': 'inlinks',
        'Is In Sitemap': 'in_sitemap',
        'Content Type': 'content_type',
        'Word Count': 'word_count',
        'Size (bytes)': 'size',
        'Response Time': 'response_time',
        'Indexability Status': 'indexability_status',
    }
    sf_rename = {k: v for k, v in SF_COL_MAP.items() if k in df_raw.columns and v not in df_raw.columns}
    if sf_rename:
        df_raw.rename(columns=sf_rename, inplace=True)
        print(f"  SF columns normalized: {list(sf_rename.keys())}")

    # Auto-detectar GSC (cuando profiler_csv no está disponible, e.g. cloud)
    if not HAS_GSC:
        HAS_GSC = (
            'impressions' in df_raw.columns and
            pd.to_numeric(df_raw['impressions'], errors='coerce').fillna(0).sum() > 0
        )
        if HAS_GSC:
            print("  GSC data auto-detected from columns")

    # Normalizar decimales antes de coerción numérica
    DECIMAL_FIX_COLS = ['ctr', 'position', 'response_time']
    if DECIMAL_COMMA:
        for col in DECIMAL_FIX_COLS:
            if col in df_raw.columns:
                df_raw[col] = df_raw[col].astype(str).str.replace(',', '.', regex=False)

    NUM_COLS = ['status', 'inlinks', 'depth', 'word_count', 'size',
                'response_time', 'impressions', 'clicks', 'ctr', 'position', 'title_len']
    for col in NUM_COLS:
        if col in df_raw.columns:
            df_raw[col] = pd.to_numeric(df_raw[col], errors='coerce').fillna(0)

    # Normalizar CTR 0-100 → 0-1 (auto-detect escala)
    if HAS_GSC and 'ctr' in df_raw.columns:
        _ctr_num = pd.to_numeric(df_raw['ctr'], errors='coerce').fillna(0)
        if _ctr_num.max() > 1:
            df_raw['ctr'] = _ctr_num / 100
            print("  CTR normalizado de escala 0-100 a 0-1")

    # Filtrar HTML
    if 'content_type' in df_raw.columns:
        df = df_raw[df_raw['content_type'].astype(str).str.contains('text/html', na=False)].copy()
    else:
        df = df_raw.copy()

    # Verificar inlinks con datos reales
    HAS_INLINKS_DATA = 'inlinks' in df.columns and df['inlinks'].sum() > 0

    print(f"  Filas raw: {len(df_raw):,} → HTML: {len(df):,}")
    print(f"  GSC: {HAS_GSC} | Inlinks: {HAS_INLINKS_DATA}")


    # ──────────────────────────────────────────────────────────────────────────────
    # 3. CLASIFICACIÓN DE URLs
    # ──────────────────────────────────────────────────────────────────────────────

    def extract_lang(url):
        """Detecta el subdirectorio de idioma de la URL."""
        if not SITE_LANGS or not isinstance(url, str):
            return 'root'
        for lang in SITE_LANGS:
            if f'/{lang}/' in url or url.rstrip('/').endswith(f'/{lang}'):
                return lang
        return 'root'


    def url_type(url):
        """Clasifica la URL según los patrones del config. First match wins."""
        if not isinstance(url, str):
            return 'other'

        # Extraer path (sin dominio ni query string)
        raw = url.replace('https://', '').replace('http://', '').split('?')[0]
        parts = raw.split('/', 1)
        path = '/' + (parts[1] if len(parts) > 1 else '')

        # Homepage
        if path in ('/', '') or re.fullmatch(r'/+', path):
            return 'homepage'

        # Sistema
        for sp in SYSTEM_PATTERNS:
            if sp in path:
                return 'system'

        # Patrones en orden (first match wins)
        for pattern, ptype in URL_PATTERNS:
            if pattern in path:
                return ptype

        return 'other'


    def is_pagination(url):
        """Detecta paginación por query param o segmento de ruta."""
        if not isinstance(url, str):
            return False
        u = url.lower()
        for param in PAGINATION_PARAMS:
            if f'{param}=' in u or f'/{param}/' in u:
                return True
        if PAGINATION_PATH_REGEX and re.search(PAGINATION_PATH_REGEX, u):
            return True
        return False


    def has_double_slash(url):
        """Detecta // en la ruta (excluye el protocolo)."""
        if not isinstance(url, str):
            return False
        path_part = url.split('://', 1)[-1]
        return '//' in path_part


    df['lang_dir']       = df['url'].apply(extract_lang)
    df['url_type']       = df['url'].apply(url_type)
    df['is_pagination']  = df['url'].apply(is_pagination)
    df['has_double_slash'] = df['url'].apply(has_double_slash)

    print(f"\nURL types: {df['url_type'].value_counts().to_dict()}")


    # ──────────────────────────────────────────────────────────────────────────────
    # 4. INDEXABILIDAD
    # ──────────────────────────────────────────────────────────────────────────────

    # Valores "indexable" positivos por idioma SF
    INDEXABLE_VALUE = {
        'es': 'Indexable', 'en': 'Indexable', 'fr': 'Indexable',
        'de': 'Indexierbar', 'it': 'Indicizzabile', 'pt': 'Indexável',
    }
    idx_pos_val = INDEXABLE_VALUE.get(SF_LANG, 'Indexable')

    if 'indexable' in df.columns:
        df_indexable    = df[df['indexable'].astype(str).str.strip() == idx_pos_val]
        df_no_indexable = df[df['indexable'].astype(str).str.strip() != idx_pos_val]
    else:
        df_indexable    = df[df['status'] == 200]
        df_no_indexable = df[df['status'] != 200]

    total_indexable    = len(df_indexable)
    total_no_indexable = len(df_no_indexable)
    print(f"  Indexable: {total_indexable:,}  /  No indexable: {total_no_indexable:,}")


    # ──────────────────────────────────────────────────────────────────────────────
    # 5. MÉTRICAS
    # ──────────────────────────────────────────────────────────────────────────────
    print("\nCalculando métricas...")

    # Status
    total_html      = len(df)
    status_200      = int((df['status'] == 200).sum())
    status_0        = int((df['status'] == 0).sum())
    status_301_count= int((df['status'] == 301).sum())
    status_404_count= int((df['status'] == 404).sum())
    status_5xx_count= int((df['status'] >= 500).sum())

    # Razones no-indexabilidad
    if 'indexability_status' in df.columns:
        idx_status = df['indexability_status'].value_counts()
        total_blocked_robots = int(idx_status[idx_status.index.str.contains('robots', case=False, na=False)].sum())
        total_canonicalized  = int(idx_status[idx_status.index.str.contains('Canoni', case=False, na=False)].sum())
        total_redirects_idx  = int(idx_status[idx_status.index.str.contains('[Rr]edir', na=False)].sum())
    else:
        total_blocked_robots = total_canonicalized = total_redirects_idx = 0

    # Error sets
    df_503 = df[df['status'] == 503]
    df_5xx = df[df['status'] >= 500]
    df_4xx = df[(df['status'] >= 400) & (df['status'] < 500)]
    df_404 = df[df['status'] == 404]
    df_301 = df[df['status'] == 301]

    # Double slash (todas las HTML, no solo indexables)
    df_double_slash_all       = df[df['has_double_slash']]
    df_double_slash_indexable = df_indexable[df_indexable['has_double_slash']]
    n_ds = len(df_double_slash_all)

    # Paginaciones
    df_paginaciones          = df[df['is_pagination']]
    df_paginaciones_indexable = df_indexable[df_indexable['is_pagination']]
    n_pag_total  = len(df_paginaciones)
    n_pag_index  = len(df_paginaciones_indexable)

    # Cart-reco (auto-detect aunque HAS_CART_RECO=False)
    _mask_cart = (
        (df_indexable['url_type'] == 'cart_reco') |
        (df_indexable['url'].str.contains('cart-recommendations', case=False, na=False))
    )
    df_cart_reco_indexable = df_indexable[_mask_cart]
    if not HAS_CART_RECO and len(df_cart_reco_indexable) > 0:
        HAS_CART_RECO = True  # auto-detectado

    # Locale especial (ej: /ca/)
    if SPECIAL_LOCALE:
        df_ca   = df[df['lang_dir'] == SPECIAL_LOCALE]
        ca_url_count = len(df_ca)
        if 'indexability_status' in df.columns:
            df_ca_blocked = df_ca[df_ca['indexability_status'].astype(str).str.contains('robots', case=False, na=False)]
        else:
            df_ca_blocked = pd.DataFrame(columns=df.columns)
    else:
        df_ca        = pd.DataFrame(columns=df.columns)
        df_ca_blocked = pd.DataFrame(columns=df.columns)
        ca_url_count = 0

    # Breakdown por idioma
    lang_indexable      = {}
    langs_with_indexable = []
    for lang in SITE_LANGS:
        n = len(df_indexable[df_indexable['lang_dir'] == lang])
        lang_indexable[lang] = n
        if n > 0:
            langs_with_indexable.append(lang)
    lang_indexable['root'] = len(df_indexable[df_indexable['lang_dir'] == 'root'])

    # On-page (páginas SEO)
    df_seo = df_indexable[df_indexable['url_type'].isin(SEO_TYPES)]

    meta_col = 'meta_desc'
    if meta_col in df.columns:
        df_no_meta = df_seo[df_seo[meta_col].isna() | (df_seo[meta_col].astype(str).str.strip() == '')]
        n_no_meta_products    = len(df_no_meta[df_no_meta['url_type'] == 'product'])
        n_no_meta_collections = len(df_no_meta[df_no_meta['url_type'].isin(['collection', 'collections_root'])])
        n_no_meta_pages       = len(df_no_meta[df_no_meta['url_type'].isin(['page', 'blog'])])
    else:
        df_no_meta = pd.DataFrame(columns=df.columns)
        n_no_meta_products = n_no_meta_collections = n_no_meta_pages = 0
    n_no_meta = len(df_no_meta)

    h1_col = 'h1'
    if h1_col in df.columns:
        df_no_h1     = df_seo[df_seo[h1_col].isna() | (df_seo[h1_col].astype(str).str.strip() == '')]
        n_no_h1_info = len(df_no_h1[df_no_h1['url_type'].isin(['page', 'blog'])])
    else:
        df_no_h1     = pd.DataFrame(columns=df.columns)
        n_no_h1_info = 0
    n_no_h1 = len(df_no_h1)

    # Titles cortos en colecciones/categorías
    df_seo_collections = df_seo[df_seo['url_type'].isin(['collection', 'collections_root'])]
    if 'title_len' in df.columns:
        df_short_title = df_seo_collections[df_seo_collections['title_len'] < T_TITLE_SHORT_CHARS]
    elif 'title' in df.columns:
        df_short_title = df_seo_collections[df_seo_collections['title'].astype(str).str.len() < T_TITLE_SHORT_CHARS]
    else:
        df_short_title = pd.DataFrame(columns=df.columns)
    n_short_title = len(df_short_title)

    # GSC
    if HAS_GSC:
        df_gsc = df[df['impressions'] > 0]
        total_impressions = int(df_gsc['impressions'].sum())
        total_clicks      = int(df_gsc['clicks'].sum())
        overall_ctr       = total_clicks / total_impressions if total_impressions > 0 else 0

        df_pos_11_20 = df_indexable[
            (df_indexable['position'] >= T_POS_OPP_MIN) &
            (df_indexable['position'] <= T_POS_OPP_MAX) &
            (df_indexable['impressions'] > 0)
        ]
        df_low_ctr = df_indexable[
            (df_indexable['ctr'] < T_CTR_LOW) &
            (df_indexable['impressions'] >= T_IMPRESSIONS_MIN_CTR) &
            (df_indexable['position'] <= 10) &
            (df_indexable['position'] > 0)
        ]
        df_impr_no_clicks = df_indexable[
            (df_indexable['impressions'] >= T_IMPRESSIONS_NO_CLICKS) &
            (df_indexable['clicks'] == 0)
        ]
        if HAS_INLINKS_DATA:
            df_seo_demand = df_indexable[
                (df_indexable['impressions'] >= T_IMPRESSIONS_DEMAND) &
                (df_indexable['inlinks'] <= T_INLINKS_LOW) &
                (df_indexable['url_type'].isin(SEO_TYPES))
            ]
        else:
            df_seo_demand = df_indexable[
                (df_indexable['impressions'] >= T_IMPRESSIONS_DEMAND) &
                (df_indexable['url_type'].isin(SEO_TYPES))
            ]
        n_demand = len(df_seo_demand)
    else:
        df_gsc = df_pos_11_20 = df_low_ctr = df_impr_no_clicks = df_seo_demand = pd.DataFrame(columns=df.columns)
        total_impressions = total_clicks = n_demand = 0
        overall_ctr = 0.0

    print(f"  Métricas calculadas. Demand: {n_demand} | Pos11-20: {len(df_pos_11_20)} | Low CTR: {len(df_low_ctr)}")

    # ── Métricas adicionales ────────────────────────────────────────────────────

    # Redirects 302 (temporales)
    df_302 = df[df['status'] == 302]

    # URLs sensibles indexables (cart, checkout, login, search…)
    _SENSITIVE = [
        '/cart', '/checkout', '/login', '/account/', '/register',
        '/buscar', '/search?', '/cesta', '/carrito', '?q=', '?s=',
        '/wishlist', '/mi-cuenta', '/my-account', '/basket',
        '/carro', '/pedido', '/factura',
    ]
    _mask_sensitive = df_indexable['url'].apply(
        lambda u: any(p in u.lower() for p in _SENSITIVE) if isinstance(u, str) else False
    )
    df_sensitive = df_indexable[_mask_sensitive]

    # Facetas/filtros indexables
    FACET_URL_PATTERNS = getattr(cfg, 'FACET_URL_PATTERNS', [])
    _default_facet_regex = (
        r'-[\d]+-f\b|[?&](filtro|filter|color|colour|marca|brand|precio|price'
        r'|sort|order|talla|size|genero|gender|material|acabado|finish)[=]'
    )
    _extra_facets = '|'.join([re.escape(p) for p in FACET_URL_PATTERNS]) if FACET_URL_PATTERNS else ''
    _facet_combined = _default_facet_regex + ('|' + _extra_facets if _extra_facets else '')
    _mask_facets = df_indexable['url'].apply(
        lambda u: bool(re.search(_facet_combined, u, re.I)) if isinstance(u, str) else False
    )
    df_facets = df_indexable[_mask_facets]

    # Canonical a URL diferente (non-self canonical en páginas indexables)
    if 'canonical' in df.columns:
        _mask_non_self_can = (
            df_indexable['canonical'].notna() &
            (df_indexable['canonical'].astype(str).str.strip() != '') &
            (
                df_indexable['canonical'].astype(str).str.strip()
                != df_indexable['url'].astype(str).str.strip()
            )
        )
        df_non_self_canonical = df_indexable[_mask_non_self_can]
    else:
        df_non_self_canonical = pd.DataFrame(columns=df.columns)

    # Thin content (páginas SEO con word_count muy bajo)
    if 'word_count' in df.columns:
        df_thin = df_indexable[
            (df_indexable['url_type'].isin(SEO_TYPES)) &
            (df_indexable['word_count'] > 0) &
            (df_indexable['word_count'] < T_WORD_COUNT_THIN)
        ]
    else:
        df_thin = pd.DataFrame(columns=df.columns)

    # Huérfanas potenciales (tráfico GSC pero 0 inlinks internos)
    if HAS_GSC and HAS_INLINKS_DATA:
        df_orphans = df_indexable[
            (df_indexable['impressions'] > T_ORPHAN_IMPRESSIONS) &
            (df_indexable['inlinks'] == 0) &
            (df_indexable['url_type'].isin(SEO_TYPES))
        ]
    else:
        df_orphans = pd.DataFrame(columns=df.columns)

    print(f"  302: {len(df_302)} | Sensibles: {len(df_sensitive)} | Facetas: {len(df_facets)}")
    print(f"  Non-self canonical: {len(df_non_self_canonical)} | Thin: {len(df_thin)} | Huérfanas: {len(df_orphans)}")

    # ─── PRE-CÓMPUTOS T21–T29 ──────────────────────────────────────────────────

    # Flags de disponibilidad de columnas
    HAS_SITEMAP_DATA   = 'in_sitemap' in df.columns
    HAS_RESPONSE_TIME  = 'response_time' in df.columns
    HAS_TITLE_1        = 'title_1' in df.columns
    HAS_TITLE_LEN      = 'title_1_length' in df.columns
    HAS_CANONICAL_COL  = 'canonical_link_element_1' in df.columns

    # T21 — URLs no-indexables incluidas en el sitemap
    if HAS_SITEMAP_DATA:
        _mask_noindex_sitemap = (
            (df['in_sitemap'] == True) &
            (df['indexability'] != 'Indexable')
        )
        df_noindex_in_sitemap = df[_mask_noindex_sitemap].copy()
    else:
        df_noindex_in_sitemap = pd.DataFrame(columns=df.columns)

    # T22 — URLs indexables ausentes del sitemap
    if HAS_SITEMAP_DATA:
        _mask_missing_sitemap = (
            (df['in_sitemap'] == False) &
            (df['indexability'] == 'Indexable') &
            (df['status'] == 200)
        )
        df_missing_from_sitemap = df[_mask_missing_sitemap].copy()
    else:
        df_missing_from_sitemap = pd.DataFrame(columns=df.columns)

    # T23 — Páginas con tiempo de respuesta >3 s (solo HTML indexables 200)
    T_SLOW_RESPONSE = 3.0
    if HAS_RESPONSE_TIME:
        df_slow = df_indexable[
            (df_indexable['response_time'] > T_SLOW_RESPONSE) &
            (df_indexable['status'] == 200)
        ].copy()
    else:
        df_slow = pd.DataFrame(columns=df.columns)

    # T24 — URLs con parámetros indexables
    df_parametered = df_indexable[
        df_indexable['url'].str.contains(r'\?', na=False)
    ].copy()

    # T25 — URLs largas (>115 caracteres) indexables
    T_URL_MAX_LEN = 115
    df_long_urls = df_indexable[
        df_indexable['url'].str.len() > T_URL_MAX_LEN
    ].copy()

    # T26 — Titles duplicados entre páginas indexables
    if HAS_TITLE_1:
        _titles_ser = df_indexable['title_1'].dropna()
        _titles_ser = _titles_ser[_titles_ser.str.strip() != '']
        _dup_titles = _titles_ser[_titles_ser.duplicated(keep=False)]
        df_dup_titles = df_indexable[
            df_indexable.index.isin(_dup_titles.index)
        ].copy()
        n_unique_dup_titles = int(df_indexable.loc[
            df_indexable.index.isin(_dup_titles.index), 'title_1'
        ].nunique())
    else:
        df_dup_titles = pd.DataFrame(columns=df.columns)
        n_unique_dup_titles = 0

    # T27 — Páginas noindex con impresiones en GSC
    if HAS_GSC:
        df_noindex_with_gsc = df[
            (df['indexability'] != 'Indexable') &
            (df['impressions'] > 0)
        ].copy()
    else:
        df_noindex_with_gsc = pd.DataFrame(columns=df.columns)

    # T28 — Titles largos (>60 chars) en páginas indexables
    if HAS_TITLE_LEN:
        df_long_titles = df_indexable[
            df_indexable['title_1_length'] > 60
        ].copy()
    else:
        df_long_titles = pd.DataFrame(columns=df.columns)

    # T29 — Canonical ausente en páginas 200 indexables
    if HAS_CANONICAL_COL:
        df_no_canonical = df_indexable[
            (df_indexable['status'] == 200) &
            (
                df_indexable['canonical_link_element_1'].isna() |
                (df_indexable['canonical_link_element_1'].astype(str).str.strip() == '')
            )
        ].copy()
    else:
        df_no_canonical = pd.DataFrame(columns=df.columns)

    print(f"  Noindex en sitemap: {len(df_noindex_in_sitemap)} | Faltan en sitemap: {len(df_missing_from_sitemap)}")
    print(f"  Lentas >3s: {len(df_slow)} | Parámetros: {len(df_parametered)} | URLs largas: {len(df_long_urls)}")
    print(f"  Titles dup: {len(df_dup_titles)} | Noindex+GSC: {len(df_noindex_with_gsc)} | Sin canonical: {len(df_no_canonical)}")

    # T30 — Sin H2 en páginas indexables SEO
    HAS_H2_COL = 'h2_1' in df.columns
    if HAS_H2_COL:
        df_no_h2 = df_indexable[
            (df_indexable['url_type'].isin(SEO_TYPES)) &
            (
                df_indexable['h2_1'].isna() |
                (df_indexable['h2_1'].astype(str).str.strip() == '')
            )
        ].copy()
    else:
        df_no_h2 = pd.DataFrame(columns=df.columns)

    # T31 — Title = H1 exacto (páginas indexables)
    HAS_H1_COL   = 'h1_1' in df.columns
    if HAS_TITLE_1 and HAS_H1_COL:
        _t1 = df_indexable['title_1'].fillna('').astype(str).str.strip().str.lower()
        _h1 = df_indexable['h1_1'].fillna('').astype(str).str.strip().str.lower()
        df_title_eq_h1 = df_indexable[
            (_t1 != '') & (_h1 != '') & (_t1 == _h1)
        ].copy()
    else:
        df_title_eq_h1 = pd.DataFrame(columns=df.columns)

    # T32 — Crawl depth alto (>4 clicks desde home) en páginas indexables con tráfico
    HAS_DEPTH_COL = 'depth' in df.columns
    T_MAX_DEPTH   = 4
    if HAS_DEPTH_COL:
        _deep_base = df_indexable[df_indexable['depth'] > T_MAX_DEPTH]
        if HAS_GSC and 'impressions' in _deep_base.columns:
            df_deep = _deep_base[_deep_base['impressions'] > 0].copy()
        else:
            df_deep = _deep_base.copy()
    else:
        df_deep = pd.DataFrame(columns=df.columns)

    print(f"  Sin H2: {len(df_no_h2)} | Title=H1: {len(df_title_eq_h1)} | Depth>{T_MAX_DEPTH}: {len(df_deep)}")


    # ──────────────────────────────────────────────────────────────────────────────
    # 6. TAREAS
    # ──────────────────────────────────────────────────────────────────────────────
    print("\nBuilding tasks...")

    tasks = []

    def add_task(tid, prioridad, categoria, tarea, desc_corta, evidencia,
                 causa_probable, que_hacer, donde_detectarlo, esfuerzo, impacto,
                 riesgo, responsable, validacion, urls_ejemplo):
        tasks.append({
            'ID': tid, 'Prioridad': prioridad, 'Categoría': categoria,
            'Tarea': tarea, 'Descripción corta': desc_corta, 'Evidencia': evidencia,
            'Causa probable': causa_probable, 'Qué hacer': que_hacer,
            'Dónde detectarlo': donde_detectarlo, 'Esfuerzo': esfuerzo,
            'Impacto': impacto, 'Riesgo': riesgo, 'Responsable': responsable,
            'Validación': validacion, 'URLs ejemplo': urls_ejemplo,
        })


    def sample_urls(df_sub, n=5, sort_by='impressions', ascending=False):
        if len(df_sub) == 0:
            return 'N/D'
        if sort_by in df_sub.columns:
            df_sub = df_sub.sort_values(sort_by, ascending=ascending)
        return '\n'.join(df_sub['url'].head(n).tolist())


    # T01 — Locale especial bloqueado (P0, condicional)
    if SPECIAL_LOCALE and ca_url_count > 0:
        locale_label = f'/{SPECIAL_LOCALE}/'
        add_task(
            'T01', 'P0', f'Indexación / {SPECIAL_LOCALE.upper()}',
            f'Resolver bloqueo de robots.txt para el locale {locale_label}',
            f'El locale {locale_label} tiene {ca_url_count:,} URLs detectadas. '
            f'robots.txt bloquea el rastreo de {locale_label} — SF no puede confirmar el estado real.',
            f'{ca_url_count:,} URL(s) con locale {locale_label} detectadas en el crawl. '
            'robots.txt contiene Disallow para este locale. '
            'SF encontró estas URLs enlazadas internamente pero no puede rastrearlas.',
            f'robots.txt incluye una regla Disallow que cubre {locale_label}. '
            'Si es intencionado: las URLs no indexarán (ni Google puede rastrearlas). '
            'Si es un error de configuración: hay contenido bloqueado involuntariamente.',
            f'1. Confirmar con el equipo si el locale {locale_label} debe indexar. '
            f'2a. Si SÍ: eliminar Disallow: {locale_label} de robots.txt. '
            'Re-crawl con SF para auditar todas las URLs del locale. '
            f'2b. Si NO: mantener el Disallow y añadir canonical en las URLs {locale_label} '
            'apuntando a su equivalente en el idioma principal.',
            f'robots.txt del sitio. SF > Configuración > Robots.txt (deshabilitar y re-crawlear).',
            'Bajo', 'Alto', 'Alto', 'SEO + Dev',
            f'Si se decide indexar: re-crawl confirma todas las URLs {locale_label} en status 200 e indexables. '
            f'Si no indexar: 0 URLs {locale_label} en GSC Cobertura.',
            sample_urls(df_ca, sort_by='impressions')
        )

    # T02 — Errores 5xx (P0, condicional)
    n_5xx = len(df_5xx)
    n_503 = len(df_503)
    if n_5xx > 0:
        add_task(
            'T02', 'P0', 'Técnico / Servidor',
            f'Resolver {n_5xx:,} errores de servidor (5xx) detectados en el crawl',
            f'{n_5xx:,} URLs devolvieron un error de servidor durante el crawl '
            f'({n_503:,} con código 503).',
            f'{n_5xx:,} URLs con status 5xx ({n_503:,} × 503). '
            'Los errores de servidor impiden el rastreo e indexación.',
            'Timeouts del servidor, errores de aplicación o bloqueo a nivel de CDN/firewall para el bot de SF.',
            '1. Verificar en producción si las URLs devuelven 5xx de forma consistente. '
            '2. Si es timeout del crawler: re-crawl en horario valle. '
            '3. Si el error es real: escalar a Dev urgente.',
            'SF > Respuesta > Códigos de respuesta > 5xx.',
            'Bajo', 'Alto', 'Alto', 'Dev + SEO',
            'Re-crawl: 0 URLs con status 5xx o 503.',
            sample_urls(df_503 if n_503 > 0 else df_5xx, sort_by='impressions')
        )

    # T03 — Errores 4xx (P0, condicional)
    n_4xx = len(df_4xx)
    n_404 = len(df_404)
    if n_4xx > 0:
        top_404_url  = df_404.nlargest(1, 'impressions').iloc[0]['url'] if n_404 > 0 else ''
        top_404_impr = int(df_404['impressions'].max()) if n_404 > 0 else 0

        # Detalle de inlinks cuando está disponible
        if HAS_INLINKS_DATA:
            df_4xx_critical = df_4xx[df_4xx['inlinks'] > T_INLINKS_4XX_CRITICAL]
            n_4xx_critical  = len(df_4xx_critical)
            inlinks_detail = (
                f' De ellas, {n_4xx_critical:,} reciben más de {T_INLINKS_4XX_CRITICAL:,} inlinks internos (prioridad crítica).'
                if n_4xx_critical > 0 else ''
            )
        else:
            df_4xx_critical = pd.DataFrame(columns=df.columns)
            n_4xx_critical  = 0
            inlinks_detail  = ''

        evid_404 = (
            f'{n_4xx:,} URLs con status 4xx (de las cuales {n_404:,} son 404).{inlinks_detail} '
            f'La 404 con más demanda GSC: {top_404_url} ({top_404_impr:,} impresiones).'
            if top_404_url else
            f'{n_4xx:,} URLs con status 4xx (de las cuales {n_404:,} son 404).{inlinks_detail}'
        )
        t03_prioridad = 'P0'
        add_task(
            'T03', t03_prioridad, 'Técnico / Errores',
            f'Corregir {n_4xx:,} URLs con error 4xx ({n_404:,} × 404)',
            f'{n_4xx:,} URLs activas en el crawl devuelven error 4xx. '
            'Pueden estar enlazadas internamente o en sitemaps.',
            evid_404,
            'URLs eliminadas sin redirect. Cambios de URL sin actualizar enlaces internos. '
            'Errores en la generación de URLs dinámicas.',
            '1. Para cada 404 con tráfico GSC: crear redirect 301 a la URL equivalente. '
            + (f'2. URGENTE: las {n_4xx_critical:,} URLs con más de {T_INLINKS_4XX_CRITICAL:,} inlinks tienen máximo impacto en PageRank — resolver primero. '
               if n_4xx_critical > 0 else
               '2. ') +
            'Identificar qué páginas enlazan a las 4xx (SF > All Inlinks). '
            '3. Actualizar los enlaces internos rotos. '
            '4. Eliminar las 4xx de sitemaps XML si están presentes.',
            'SF > Respuesta > Códigos de respuesta > 4xx. GSC > Cobertura > 404.',
            'Medio', 'Alto', 'Alto', 'Dev + SEO',
            'Re-crawl: 0 URLs enlazadas internamente con status 4xx. '
            'GSC Cobertura: reducción de errores 404.',
            sample_urls(
                df_4xx_critical if n_4xx_critical > 0 else df_404,
                sort_by='inlinks' if HAS_INLINKS_DATA else 'impressions'
            )
        )

    # T04 — Redirects 301 (P1, condicional)
    n_301 = len(df_301)
    if n_301 > 0:
        if HAS_INLINKS_DATA:
            df_301_inlinks = df_301[df_301['inlinks'] > 0]
            n_301_inlinks  = len(df_301_inlinks)
            evid_301 = (
                f'{n_301:,} URLs con redirect 301 detectadas en el crawl. '
                f'{n_301_inlinks:,} tienen inlinks internos (actualizar el enlace al destino final). '
                'Los redirects encadenados diluyen PageRank y ralentizan el rastreo.'
            )
            valid_301 = (
                'Re-crawl: 0 URLs con status 301 reciben inlinks internos. '
                'SF > All Inlinks: los enlaces apuntan directamente a las URLs destino.'
            )
        else:
            evid_301 = (
                f'{n_301:,} URLs con redirect 301. '
                'Para ver qué páginas enlazan a estas 301s: usar SF > All Inlinks export.'
            )
            valid_301 = (
                'Re-crawl: 0 URLs con status 301 reciben inlinks internos. '
                'Verificar con SF > All Inlinks export.'
            )
        add_task(
            'T04', 'P1', 'Técnico / Redirects',
            f'Actualizar los {n_301:,} enlaces internos que apuntan a redirects 301',
            f'{n_301:,} URLs indexadas como 301. '
            'Los enlaces internos deben apuntar a la URL final para no perder autoridad.',
            evid_301,
            'Templates y CMS generan enlaces a URLs antiguas que ya redirigen. '
            'Actualizaciones de producto/categoría sin actualizar referencias.',
            '1. Exportar SF > All Inlinks y filtrar por Destination Status = 301. '
            '2. Priorizar las 301s con más inlinks entrantes. '
            '3. Actualizar los enlaces en los templates para apuntar a la URL destino. '
            f'4. En {PLATFORM}: revisar menús, breadcrumbs y templates de navegación.',
            'SF > Respuesta > Redirecciones > 301. SF > All Inlinks.',
            'Medio', 'Alto', 'Bajo', 'Dev + SEO',
            valid_301,
            sample_urls(df_301, sort_by='impressions')
        )

    # T05 — Double slash (P1, condicional)
    if n_ds > 0:
        n_ds_idx = len(df_double_slash_indexable)
        add_task(
            'T05', 'P1', 'Técnico / URLs',
            f'Corregir {n_ds:,} URLs con doble slash (//) en la ruta',
            f'{n_ds:,} URLs detectadas con "//" en su ruta '
            f'({n_ds_idx:,} indexables, el resto canonicalizadas) — posible contenido duplicado.',
            f'{n_ds:,} URLs con // en la ruta: {n_ds_idx:,} indexables + '
            f'{n_ds - n_ds_idx:,} canonicalizadas. '
            'Google puede tratar la versión con // y sin // como URLs distintas → duplicados.',
            f'Error en los templates de {PLATFORM}: concatenación incorrecta de variables de ruta.',
            '1. Identificar el template que genera las URLs con //. '
            '2. Corregir la concatenación en el código. '
            '3. Añadir redirect 301 de la versión con // a la versión normalizada.',
            'SF > All URLs > filtrar columna URL por //.',
            'Bajo', 'Medio', 'Bajo', 'Dev',
            'Re-crawl: 0 URLs con // en la ruta.',
            sample_urls(df_double_slash_all, sort_by='inlinks')
        )

    # T06 — Cart-reco indexable (P1, Shopify-específico)
    if HAS_CART_RECO and len(df_cart_reco_indexable) > 0:
        n_cart = len(df_cart_reco_indexable)
        add_task(
            'T06', 'P1', f'Indexación / {PLATFORM}',
            f'Bloquear la indexación de {n_cart:,} URLs de cart-recommendations',
            f'{n_cart:,} URLs de tipo cart-recommendations son indexables. '
            'Estas páginas técnicas no deben aparecer en Google.',
            f'{n_cart:,} URLs con "cart-recommendations" indexadas en el crawl.',
            f'Las URLs de cart-recommendations son endpoints técnicos de {PLATFORM} '
            'para sugerencias de productos. No tienen contenido editorial propio.',
            '1. Añadir meta robots noindex en el template liquid que genera estas páginas. '
            '2. Alternativamente, añadir Disallow: /recommendations/ en robots.txt.',
            f'SF > Indexabilidad = Indexable > filtrar URL contiene "cart-recommendations".',
            'Bajo', 'Medio', 'Bajo', 'Dev',
            'Re-crawl: 0 URLs de cart-recommendations con estado Indexable.',
            sample_urls(df_cart_reco_indexable)
        )

    # T07 — Paginaciones indexables (P1, condicional)
    if n_pag_index > 0:
        gsc_pag_count = len(df_paginaciones_indexable[df_paginaciones_indexable['impressions'] > 0]) if HAS_GSC else 0
        add_task(
            'T07', 'P1', 'Indexación / Paginación',
            f'Revisar {n_pag_index:,} páginas de paginación indexables',
            f'{n_pag_index:,} URLs de paginación son indexables. '
            'Las paginaciones raramente deben indexarse ya que diluyen autoridad.',
            f'{n_pag_index:,} URLs de paginación indexadas (de {n_pag_total:,} detectadas). '
            + (f'Con señal GSC activa: {gsc_pag_count} URLs.' if gsc_pag_count > 0 else ''),
            'Ausencia de meta robots canonical o noindex en páginas de paginación.',
            '1. Añadir canonical en /page/2, /page/3... apuntando a la página /1. '
            '2. Evaluar noindex para page/2 en adelante. '
            '3. Excluir paginaciones de los sitemaps XML.',
            'SF > filtrar URL contiene "page=" o "/page/". Columna Indexabilidad.',
            'Bajo', 'Medio', 'Bajo', 'Dev + SEO',
            'Re-crawl: 0 URLs de paginación indexadas sin canonical a la URL base.',
            sample_urls(df_paginaciones_indexable, sort_by='impressions')
        )

    # T08 — Meta descriptions (P1, condicional)
    if n_no_meta > 0:
        add_task(
            'T08', 'P1', 'Metadatos / Meta Description',
            f'Añadir meta description a las {n_no_meta:,} páginas SEO que carecen de ella',
            f'{n_no_meta:,} páginas SEO indexables no tienen meta description. '
            'Google puede generar snippets poco atractivos que reducen el CTR.',
            f'{n_no_meta:,} páginas SEO sin meta description: '
            f'{n_no_meta_products:,} productos, {n_no_meta_collections:,} colecciones/categorías, '
            f'{n_no_meta_pages:,} páginas estáticas.',
            f'Los templates de {PLATFORM} no tienen configurada la meta description por defecto. '
            'Las páginas creadas sin rellenar el campo aparecen sin meta desc.',
            '1. Implementar meta description dinámica por defecto en el template '
            '(ej: [Nombre del producto] — Comprar online. Envío rápido | [Marca]). '
            '2. Priorizar las URLs con más impresiones en GSC. '
            '3. Completar manualmente las meta descriptions de las 20-30 URLs de mayor tráfico.',
            'SF > Meta Description > Ausente + filtrar Indexabilidad = Indexable.',
            'Medio', 'Alto', 'Bajo', 'SEO',
            'Re-crawl: 0 páginas SEO indexadas sin meta description. '
            'GSC: mejora de CTR medio en las URLs actualizadas.',
            sample_urls(df_no_meta, sort_by='impressions')
        )

    # T09 — H1 (P1, condicional)
    if n_no_h1 > 0:
        add_task(
            'T09', 'P1', 'Metadatos / H1',
            f'Añadir H1 a las {n_no_h1:,} páginas SEO sin heading principal',
            f'{n_no_h1:,} páginas SEO indexables carecen de H1. '
            'El H1 es la señal semántica principal del contenido para Google.',
            f'{n_no_h1:,} páginas SEO sin H1: {n_no_h1_info:,} páginas/posts.',
            f'Los templates de {PLATFORM} pueden usar elementos visuales como "título" '
            'sin implementar el tag H1 en el HTML.',
            '1. Auditar el template de cada tipo de página afectada. '
            '2. Asegurarse de que el título principal está envuelto en <h1>. '
            '3. Solo puede haber un H1 por página. '
            '4. El H1 debe incluir la keyword principal.',
            'SF > H1 > Ausente + filtrar Indexabilidad = Indexable.',
            'Bajo', 'Medio', 'Bajo', 'Dev',
            'Re-crawl: 0 páginas SEO indexadas sin H1.',
            sample_urls(df_no_h1, sort_by='impressions')
        )

    # T10 — Interlinking (P2, requiere GSC)
    if HAS_GSC and n_demand > 0:
        if HAS_INLINKS_DATA:
            desc_t10  = (
                f'{n_demand:,} páginas SEO tienen ≥{T_IMPRESSIONS_DEMAND:,} impresiones en GSC '
                f'pero solo 0-{T_INLINKS_LOW} inlinks internos. '
                'Reforzar su enlazado puede mejorar posición y tráfico directamente.'
            )
            evid_t10  = (
                f'{n_demand:,} URLs indexables con ≥{T_IMPRESSIONS_DEMAND:,} impresiones y '
                f'≤{T_INLINKS_LOW} inlinks. '
                f'Total impresiones acumuladas: {int(df_seo_demand["impressions"].sum()):,}.'
            )
            hacer_t10 = (
                '1. Exportar el listado de estas URLs (SF + GSC cruzado). '
                '2. Para cada URL: identificar 3-5 páginas relacionadas con autoridad '
                '(homepage, categorías) que puedan enlazarla con anchor text descriptivo. '
                '3. Añadir los enlaces desde los templates o en el contenido. '
                '4. Usar la estructura de categorías y el menú para distribuir autoridad. '
                '5. Considerar una sección de "más buscado" en la homepage con las 5-10 URLs top.'
            )
        else:
            desc_t10  = (
                f'{n_demand:,} páginas SEO tienen ≥{T_IMPRESSIONS_DEMAND:,} impresiones en GSC. '
                'Revisar su enlazado interno puede mejorar posición y tráfico.'
            )
            evid_t10  = (
                f'{n_demand:,} URLs indexables con ≥{T_IMPRESSIONS_DEMAND:,} impresiones. '
                f'Total impresiones: {int(df_seo_demand["impressions"].sum()):,}. '
                'No hay datos de inlinks en este export; usar SF > All Inlinks para priorizar.'
            )
            hacer_t10 = (
                '1. Exportar SF > All Inlinks y cruzar con estas URLs. '
                '2. Para las URLs con menos inlinks: añadir enlaces desde páginas con autoridad. '
                '3. Considerar una sección de "más buscado" en la homepage.'
            )
        t10_prio = 'P1' if n_demand > 500 else 'P2'
        add_task(
            'T10', t10_prio, 'Enlazado Interno',
            f'Reforzar el enlazado interno de {n_demand:,} páginas con demanda GSC',
            desc_t10, evid_t10,
            'Páginas relevantes para SEO que no reciben suficiente autoridad interna. '
            'El enlazado interno no responde a la demanda real de búsqueda.',
            hacer_t10,
            "GSC > Rendimiento > Páginas > ordenar por Impresiones. Cruzar con SF 'All Inlinks'.",
            'Medio', 'Alto', 'Bajo', 'SEO + Dev',
            'Recrawl: las URLs objetivo tienen ≥5 inlinks internos. '
            'GSC: mejora de posición media y clics en las URLs reforzadas (evaluar a 30-60 días).',
            sample_urls(df_seo_demand.sort_values('impressions', ascending=False))
        )

    # T11 — Posición 11-20 (P1, requiere GSC)
    if HAS_GSC and len(df_pos_11_20) > 0:
        n_11_20       = len(df_pos_11_20)
        top_11_20_url = df_pos_11_20.nlargest(1, 'impressions').iloc[0]['url']
        top_11_20_impr = int(df_pos_11_20['impressions'].max())
        add_task(
            'T11', 'P1', 'GSC / Crecimiento',
            'Empujar a página 1 las URLs en posición 11-20 con mayor demanda',
            f'{n_11_20:,} URLs indexables aparecen en posición 11-20 en Google. '
            'Un empuje a página 1 puede multiplicar los clics por 5-10x.',
            f'{n_11_20:,} URLs en pos {T_POS_OPP_MIN}-{T_POS_OPP_MAX} con impresiones. '
            f'La de mayor volumen: {top_11_20_url} ({top_11_20_impr:,} impresiones). '
            f'Suma de impresiones del grupo: {int(df_pos_11_20["impressions"].sum()):,}.',
            'Contenido indexado pero insuficientemente optimizado (thin content, falta de señales '
            'de autoridad, enlazado interno escaso) para competir en el top 10.',
            '1. Para las 10 URLs con más impresiones: auditar contenido (longitud, keywords en H2, '
            'imágenes con alt text, schema). '
            '2. Reforzar enlazado interno desde páginas con autoridad. '
            '3. Revisar intención de búsqueda: ¿el contenido responde exactamente a lo que busca el usuario? '
            '4. Revisar que el snippet (title + meta) es competitivo vs las URLs en top 5.',
            f'GSC > Rendimiento > Páginas. Filtrar posición {T_POS_OPP_MIN}-{T_POS_OPP_MAX}, ordenar por Impresiones desc.',
            'Alto', 'Alto', 'Bajo', 'SEO',
            'GSC Rendimiento en 30-60 días: aumento de posición media y clics en las URLs trabajadas. '
            'Al menos 5 de las 10 URLs trabajadas deben aparecer en posición ≤10.',
            sample_urls(df_pos_11_20.sort_values('impressions', ascending=False))
        )

    # T12 — CTR bajo en top 10 (P1, requiere GSC)
    if HAS_GSC and len(df_low_ctr) > 0:
        n_low_ctr = len(df_low_ctr)
        add_task(
            'T12', 'P1', 'GSC / Snippet',
            f'Mejorar el CTR de {n_low_ctr:,} URLs en top 10 con tasa de clics inferior al {T_CTR_LOW*100:.0f}%',
            f'{n_low_ctr:,} URLs aparecen en los 10 primeros resultados pero tienen CTR <{T_CTR_LOW*100:.0f}%. '
            'Son impresiones que se muestran pero no se convierten en visitas.',
            f'{n_low_ctr:,} URLs con posición ≤10, CTR <{T_CTR_LOW*100:.0f}% y '
            f'≥{T_IMPRESSIONS_MIN_CTR} impresiones. '
            f'Suma de impresiones desaprovechadas: {int(df_low_ctr["impressions"].sum()):,}. '
            f'Llevando el CTR al {T_CTR_LOW*200:.0f}% se obtendría un incremento estimado de '
            f'{int(df_low_ctr["impressions"].sum() * T_CTR_LOW):,} clics/periodo.',
            'Titles genéricos sin diferenciadores. Meta descriptions ausentes o sin CTA. '
            'Sin datos estructurados (estrellas, precio) que enriquezcan el snippet.',
            '1. Reescribir titles de las 20 URLs con más impresiones: keyword al inicio, '
            'diferenciador (precio, envío gratis) y máximo 60 chars. '
            '2. Añadir/mejorar meta descriptions con CTA directo. '
            '3. Implementar schema de Producto (precio, disponibilidad, valoraciones). '
            '4. Testear variantes de title con GSC durante 30 días.',
            f'GSC > Rendimiento > Páginas. Filtrar posición ≤10 y CTR <{T_CTR_LOW*100:.0f}%, '
            'ordenar por Impresiones desc.',
            'Medio', 'Alto', 'Bajo', 'SEO',
            'GSC Rendimiento: CTR medio del grupo objetivo supera el doble del valor actual '
            'en los 60 días posteriores al cambio.',
            sample_urls(df_low_ctr.sort_values('impressions', ascending=False))
        )

    # T13 — Titles cortos en colecciones/categorías (P2, condicional)
    if n_short_title > 0:
        type_label = 'colecciones' if PLATFORM == 'Shopify' else 'categorías/páginas'
        add_task(
            'T13', 'P2', 'Metadatos / Titles',
            f'Revisar {n_short_title:,} {type_label} con title corto o genérico',
            f'{n_short_title:,} {type_label} tienen title de menos de {T_TITLE_SHORT_CHARS} caracteres. '
            'Los titles cortos suelen ser genéricos con menor capacidad de rankear.',
            f'{n_short_title:,} {type_label} indexadas con title <{T_TITLE_SHORT_CHARS} chars.',
            f'Los templates de {PLATFORM} usan el nombre corto de la categoría como title '
            'sin añadir modificadores SEO.',
            f'1. Para las 15-20 {type_label} más visitadas: reescribir el title incluyendo '
            'keyword principal + modificador (comprar, online) + | {DOMAIN}. '
            f'2. Ejemplo: "Nike" → "Comprar Zapatillas Nike Online | {DOMAIN}". '
            '3. Fórmula: [Keyword principal] [modificador] [año si aplica] | Marca.',
            f'SF > Titles > filtrar Longitud < {T_TITLE_SHORT_CHARS} + Indexabilidad = Indexable.',
            'Bajo', 'Medio', 'Bajo', 'SEO',
            f'Re-crawl: 0 {type_label} indexadas con title <{T_TITLE_SHORT_CHARS} chars.',
            sample_urls(df_short_title.sort_values('impressions', ascending=False) if HAS_GSC and 'impressions' in df_short_title.columns else df_short_title)
        )

    # T14 — Hreflang (P2, solo si multilingual)
    if IS_MULTILINGUAL and SITE_LANGS:
        langs_str  = ', '.join([f'/{l}/' for l in SITE_LANGS])
        n_langs    = len(SITE_LANGS) + 1
        add_task(
            'T14', 'P2', 'Internacional / Hreflang',
            f'Auditar implementación de hreflang para los {n_langs} idiomas',
            'El export de SF no incluye columna de hreflang. Es necesario un crawl específico '
            'para validar los alternates de todos los idiomas.',
            f'{n_langs} versiones de idioma detectadas: {langs_str} + raíz. '
            'No se puede validar hreflang desde el CSV actual. '
            'GSC > Internacional > Errores de hreflang puede mostrar problemas existentes.',
            'El export no incluye la pestaña Hreflang de SF (requiere activar el análisis). '
            'Posibles problemas: falta de x-default, idiomas sin self-referential alternate, '
            'pares no recíprocos.',
            '1. En SF: activar Crawl > Configuration > Spider > Extraction > soporte hreflang. '
            '2. Re-crawlear el sitio y exportar la pestaña Hreflang > All. '
            '3. Verificar: a) cada idioma tiene alternate a sí mismo, b) existe x-default, '
            'c) los pares son recíprocos. '
            '4. Validar en GSC > Configuración > Internacional > Errores de hreflang.',
            'SF > Hreflang > All (requiere re-crawl). '
            'GSC > Configuración > Internacional > Errores de hreflang.',
            'Medio', 'Alto', 'Medio', 'SEO + Dev',
            'SF Hreflang: 0 errores de conflicto, URL no rastreable o falta x-default. '
            'GSC Internacional: 0 errores para las versiones principales.',
            f'https://{DOMAIN}/\n' + '\n'.join([f'https://{DOMAIN}/{l}/' for l in SITE_LANGS])
        )

    # T15 — URLs sensibles indexables (P1, condicional)
    n_sensitive = len(df_sensitive)
    if n_sensitive > 0:
        add_task(
            'T15', 'P1', 'Indexación / Seguridad',
            f'Bloquear indexación de {n_sensitive:,} URLs sensibles (carrito, checkout, login…)',
            f'{n_sensitive:,} URLs de procesos internos o privados son indexables. '
            'Estas páginas no deben aparecer en Google: no aporten valor y exponen lógica interna.',
            f'{n_sensitive:,} URLs con patrones sensibles (cart, checkout, login, account, '
            'search, buscar, cesta…) detectadas como indexables en el crawl.',
            f'Los templates de {PLATFORM} no añaden meta robots noindex por defecto '
            'en estas rutas funcionales. Pueden estar en sitemaps y recibir inlinks internos.',
            '1. Añadir <meta name="robots" content="noindex,nofollow"> en los templates '
            'de carrito, checkout, login, cuenta y búsqueda. '
            '2. Añadir Disallow en robots.txt para las rutas /cart/, /checkout/, /login/, '
            '/account/, /search/. '
            '3. Eliminar estas URLs de sitemaps XML si están presentes. '
            '4. Actualizar los enlaces internos para que no apunten a versiones indexables.',
            'SF > Indexabilidad = Indexable > filtrar URL por cart/checkout/login/account.',
            'Bajo', 'Medio', 'Bajo', 'Dev',
            'Re-crawl: 0 URLs sensibles con estado Indexable. '
            'GSC Cobertura: estas URLs desaparecen del índice en 2-4 semanas.',
            sample_urls(df_sensitive, sort_by='impressions' if HAS_GSC else 'inlinks')
        )

    # T16 — Facetas/filtros indexables (P1, condicional)
    n_facets = len(df_facets)
    if n_facets > 0:
        add_task(
            'T16', 'P1', 'Indexación / Facetas',
            f'Añadir noindex + canonical a {n_facets:,} URLs de facetas/filtros indexables',
            f'{n_facets:,} URLs de navegación facetada (filtros por color, marca, precio, talla…) '
            'son indexables. Generan contenido duplicado y diluyen autoridad.',
            f'{n_facets:,} URLs con patrones de facetas (-f, ?filtro=, ?color=, ?marca=, ?sort=…) '
            'detectadas como indexables. Cada combinación de filtros crea una URL separada '
            'con contenido prácticamente igual al de la categoría base.',
            f'El CMS ({PLATFORM}) no bloquea por defecto la indexación de URLs generadas '
            'por facetas de navegación. Con N filtros y M valores, se generan N×M URLs.',
            '1. Añadir <meta name="robots" content="noindex"> en el template de resultados '
            'de filtrado. '
            '2. Añadir canonical en cada URL de faceta apuntando a la categoría base. '
            '3. Excluir los patrones de facetas del sitemap XML. '
            '4. Si algún filtro tiene volumen de búsqueda propio (ej: "zapatillas rojas"), '
            'crear una landing page dedicada en lugar de usar la URL de faceta.',
            'SF > All URLs > filtrar URL por patrón -f / ?filtro= / ?color=.',
            'Bajo', 'Alto', 'Bajo', 'Dev + SEO',
            'Re-crawl: 0 URLs de facetas con estado Indexable. '
            'GSC Cobertura: reducción de URLs descubiertas mediante exploración.',
            sample_urls(df_facets, sort_by='impressions' if HAS_GSC else 'inlinks')
        )

    # T17 — Canonical a URL diferente (P1, condicional)
    n_non_self_can = len(df_non_self_canonical)
    if n_non_self_can > 0:
        top_can_url    = df_non_self_canonical.iloc[0]['url'] if n_non_self_can > 0 else ''
        top_can_target = df_non_self_canonical.iloc[0]['canonical'] if n_non_self_can > 0 else ''
        add_task(
            'T17', 'P1', 'Canonicals / Indexación',
            f'Revisar {n_non_self_can:,} páginas indexables con canonical apuntando a otra URL',
            f'{n_non_self_can:,} páginas tienen canonical tag configurado hacia una URL diferente '
            'a sí mismas. Puede ser intencional (variantes de producto) o un error.',
            f'{n_non_self_can:,} URLs con canonical ≠ propia URL. '
            f'Ejemplo: {top_can_url} → canonical: {top_can_target}. '
            'Si el canonical es incorrecto, Google indexará la URL equivocada.',
            'Configuración de variantes de producto que apuntan al producto padre. '
            'Error en templates: canonical hardcodeado o calculado incorrectamente. '
            'Migración de URLs sin actualizar los canonicals.',
            '1. Exportar SF > Canonicals > Contains a canonical URL e identificar patrones. '
            '2. Separar los canonicals intencionales (variantes → padre) de los accidentales. '
            '3. Para cada canonical incorrecto: corregir el tag en el template. '
            '4. Si hay variantes legítimas: asegurarse de que la URL canónica es efectivamente '
            'la versión principal y está bien optimizada.',
            'SF > Canonicals > Contains a canonical URL (filtrar Indexabilidad = Indexable).',
            'Medio', 'Alto', 'Medio', 'SEO + Dev',
            'Re-crawl: las URLs intencionales mantienen canonical correcto. '
            'Las URLs con canonical erróneo tienen self-canonical. '
            'GSC: las URLs canónicas declaradas aparecen indexadas correctamente.',
            sample_urls(df_non_self_canonical, sort_by='impressions' if HAS_GSC else 'inlinks')
        )

    # T18 — Thin content (P2, condicional)
    n_thin = len(df_thin)
    if n_thin > 0:
        avg_thin_wc = int(df_thin['word_count'].mean()) if 'word_count' in df_thin.columns else 0
        add_task(
            'T18', 'P2', 'Contenido / Calidad',
            f'Ampliar o eliminar {n_thin:,} páginas SEO con contenido escaso '
            f'(< {T_WORD_COUNT_THIN} palabras)',
            f'{n_thin:,} páginas SEO indexables tienen menos de {T_WORD_COUNT_THIN} palabras. '
            'Google considera este tipo de contenido "thin" y puede penalizarlo.',
            f'{n_thin:,} páginas SEO (productos/colecciones/blog) con menos de '
            f'{T_WORD_COUNT_THIN} palabras (media: {avg_thin_wc} palabras). '
            'Son páginas que rankearán con dificultad y pueden dañar la percepción de calidad del dominio.',
            'Templates vacíos o sin rellenar. Páginas creadas automáticamente sin contenido editorial. '
            'Descripción de producto omitida. Posts de blog en borrador publicados accidentalmente.',
            f'1. Priorizar las {min(20, n_thin)} páginas con más impresiones en GSC. '
            '2. Para productos: añadir descripción detallada (>150 palabras): '
            'características, materiales, uso, beneficios. '
            '3. Para categorías: añadir texto introductorio SEO de 100-200 palabras al inicio y al final. '
            '4. Para posts de blog vacíos: completar o poner en noindex/borrador. '
            '5. Auditar el template para añadir campos de descripción obligatorios.',
            f'SF > All URLs > filtrar Word Count < {T_WORD_COUNT_THIN} + Indexabilidad = Indexable.',
            'Alto', 'Medio', 'Bajo', 'SEO + Contenido',
            f'Re-crawl: 0 páginas SEO con word_count < {T_WORD_COUNT_THIN}. '
            'GSC: mejora de posición en las páginas ampliadas (evaluar a 30-60 días).',
            sample_urls(df_thin, sort_by='impressions' if HAS_GSC else 'word_count')
        )

    # T19 — Redirects 302 (P1, condicional)
    n_302 = len(df_302)
    if n_302 > 0:
        add_task(
            'T19', 'P1', 'Técnico / Redirects',
            f'Convertir {n_302:,} redirects 302 (temporales) en 301 permanentes',
            f'{n_302:,} URLs responden con código 302 (redirect temporal). '
            'Los 302 no transfieren autoridad de forma fiable y confunden a Google.',
            f'{n_302:,} redirects 302 detectados en el crawl. '
            'A diferencia de los 301, los 302 no trasladan PageRank de forma garantizada '
            'y Google puede indexar tanto la URL de origen como la de destino.',
            f'Configuración errónea en {PLATFORM} o en el servidor web/CDN. '
            'Uso de redirecciones temporales para casos que son permanentes '
            '(ej: 302 de la home sin www → con www).',
            '1. Identificar el destino de cada 302 (SF > Response Codes > 3xx). '
            '2. Si la redirección es permanente: cambiar 302 → 301 en el servidor/CMS. '
            '3. Prestar especial atención a 302s hacia la homepage (muy común en '
            'configs de dominio y casos de canonicalización de homepage). '
            '4. Si algún 302 es genuinamente temporal (test A/B): mantenerlo y documentarlo.',
            'SF > Respuesta > Códigos de respuesta > 3xx > filtrar por código 302.',
            'Bajo', 'Medio', 'Bajo', 'Dev',
            'Re-crawl: 0 redirects 302 (salvo los genuinamente temporales documentados). '
            'GSC: mejora de cobertura en las URLs que eran origen de 302.',
            sample_urls(df_302, sort_by='inlinks' if HAS_INLINKS_DATA else 'url')
        )

    # T20 — Huérfanas potenciales (P2, requiere GSC + inlinks)
    n_orphans = len(df_orphans)
    if n_orphans > 0:
        top_orphan_url  = df_orphans.nlargest(1, 'impressions').iloc[0]['url']
        top_orphan_impr = int(df_orphans['impressions'].max())
        add_task(
            'T20', 'P2', 'Enlazado Interno / Huérfanas',
            f'Enlazar {n_orphans:,} páginas huérfanas con tráfico GSC pero sin inlinks internos',
            f'{n_orphans:,} páginas SEO tienen impresiones en Google (>{T_ORPHAN_IMPRESSIONS:,}) '
            'pero 0 inlinks internos. Google las encuentra pero el sitio no las enlaza.',
            f'{n_orphans:,} URLs con más de {T_ORPHAN_IMPRESSIONS:,} impresiones en GSC '
            f'y 0 inlinks internos en SF. '
            f'La de mayor volumen: {top_orphan_url} ({top_orphan_impr:,} impresiones). '
            '0 inlinks = Google trabaja solo para encontrarlas, sin apoyo de la arquitectura.',
            'Páginas creadas y olvidadas sin actualizar la navegación. '
            'Productos o posts sin categoría asignada. '
            'URLs antiguas que sobrevivieron a una migración sin preservar el enlazado.',
            '1. Exportar este listado y priorizar por impresiones. '
            '2. Para las 10-15 URLs con más impresiones: identificar 2-3 páginas '
            'con autoridad (categorías padre, homepage, posts relacionados) que puedan enlazarlas. '
            '3. Añadir los enlaces con anchor text descriptivo y relevante. '
            '4. Si la URL no debería existir (contenido obsoleto): redirigir o añadir noindex.',
            "SF > All URLs > filtrar Unique Inlinks = 0 + Indexabilidad = Indexable. "
            "Cruzar con GSC > Rendimiento > Páginas ordenado por Impresiones.",
            'Bajo', 'Alto', 'Bajo', 'SEO + Dev',
            f'Re-crawl + 30 días: las URLs objetivo tienen ≥3 inlinks internos. '
            'GSC: mejora de posición media en las URLs enlazadas.',
            sample_urls(df_orphans.sort_values('impressions', ascending=False))
        )

    # T21 — URLs no-indexables en el sitemap XML (P1)
    n_noindex_sitemap = len(df_noindex_in_sitemap)
    if HAS_SITEMAP_DATA and n_noindex_sitemap > 0:
        reasons = df_noindex_in_sitemap['indexability_status'].value_counts().head(3).to_dict() if 'indexability_status' in df_noindex_in_sitemap.columns else {}
        reasons_str = ', '.join(f'{v}× {k}' for k, v in reasons.items()) if reasons else 'diversas razones'
        add_task(
            'T21', 'P1', 'Sitemaps / Inconsistencia con indexación',
            f'Eliminar {n_noindex_sitemap:,} URLs no-indexables del sitemap XML',
            f'El sitemap incluye {n_noindex_sitemap:,} URLs que Google no puede indexar '
            f'(noindex, canonicalizada a otra URL, etc.). '
            'Diluye el crawl budget y confunde a los bots.',
            f'{n_noindex_sitemap:,} URLs están en el sitemap pero marcadas como no-indexables en SF. '
            f'Razones principales: {reasons_str}. '
            'Google prioriza el sitemap para descubrir URLs indexables; incluir no-indexables '
            'genera señales contradictorias entre el sitemap y las directivas on-page.',
            'Sitemap generado automáticamente sin filtrar por indexabilidad. '
            'Cambios de indexabilidad posteriores a la generación del sitemap. '
            'Plugin de sitemap sin configuración avanzada.',
            '1. Filtrar el sitemap para incluir solo URLs status 200 + Indexable. '
            '2. Si usas un plugin (Yoast, Rank Math…): excluir noindex/canonicalizadas. '
            '3. Regenerar el sitemap y reenviar en GSC > Sitemaps. '
            '4. Verificar en GSC > Cobertura que desaparecen las URLs excluidas.',
            'SF > Sitemaps > Export all crawled in sitemaps. Filtrar Indexability ≠ Indexable.',
            'Bajo', 'Medio', 'Bajo', 'SEO + Dev',
            'Re-crawl: 0 URLs no-indexables en sitemap. GSC Cobertura: caída en "Excluida - noindex".',
            sample_urls(df_noindex_in_sitemap, sort_by='inlinks' if HAS_INLINKS_DATA else 'url')
        )

    # T22 — URLs indexables ausentes del sitemap (P1)
    n_missing_sitemap = len(df_missing_from_sitemap)
    if HAS_SITEMAP_DATA and n_missing_sitemap > 0:
        add_task(
            'T22', 'P1', 'Sitemaps / Cobertura incompleta',
            f'Añadir {n_missing_sitemap:,} URLs indexables al sitemap XML',
            f'{n_missing_sitemap:,} páginas indexables (status 200) no están en el sitemap. '
            'Google puede tardar más en descubrirlas y recrawlearlas.',
            f'{n_missing_sitemap:,} URLs con status 200 e indexabilidad "Indexable" ausentes del sitemap. '
            'El sitemap actúa como señal de prioridad de rastreo; las URLs fuera de él '
            'dependen exclusivamente del enlazado interno para su descubrimiento.',
            'Sitemap configurado con exclusiones excesivas. '
            'Nuevas secciones del sitio no añadidas al generador de sitemaps. '
            'Sitemap estático no actualizado.',
            '1. Identificar qué tipos de URL faltan (categorías, posts, productos…). '
            '2. Actualizar la configuración del generador de sitemap para incluirlas. '
            '3. Si el sitemap es estático: añadir las URLs manualmente o migrar a generación dinámica. '
            '4. Regenerar y reenviar en GSC > Sitemaps.',
            'SF > Sitemaps > In Sitemap = No. Filtrar Indexability = Indexable + Status 200.',
            'Bajo', 'Medio', 'Bajo', 'SEO + CMS',
            'Re-crawl: todas las URLs indexables aparecen en el sitemap. '
            'GSC > Sitemaps: aumento en el recuento de URLs enviadas.',
            sample_urls(df_missing_from_sitemap, sort_by='inlinks' if HAS_INLINKS_DATA else 'url')
        )

    # T23 — Tiempo de respuesta lento >3 s (P1)
    n_slow = len(df_slow)
    if HAS_RESPONSE_TIME and n_slow > 0:
        avg_slow = round(df_slow['response_time'].mean(), 2) if 'response_time' in df_slow.columns else 0
        top_slow_url = df_slow.sort_values('response_time', ascending=False).iloc[0]['url'] if n_slow > 0 else ''
        top_slow_time = round(df_slow['response_time'].max(), 2) if n_slow > 0 else 0
        add_task(
            'T23', 'P1', 'Técnico / Velocidad',
            f'Optimizar tiempo de respuesta en {n_slow:,} páginas lentas (>{T_SLOW_RESPONSE}s)',
            f'{n_slow:,} páginas indexables superan los {T_SLOW_RESPONSE}s de TTFB durante el crawl. '
            f'Media: {avg_slow}s. Máximo: {top_slow_time}s. Impacta Core Web Vitals y crawl budget.',
            f'{n_slow:,} páginas con response_time >{T_SLOW_RESPONSE}s en SF. '
            f'Tiempo medio de las afectadas: {avg_slow}s. '
            f'URL más lenta: {top_slow_url} ({top_slow_time}s). '
            'Googlebot reduce la frecuencia de rastreo en sitios con TTFB elevado; '
            'además, un TTFB alto es el principal factor del LCP.',
            'Consultas de base de datos sin índices. '
            'Plugins pesados o conflictos de caché. '
            'Hosting infradimensionado. '
            'Procesos de terceros bloqueantes en server-side.',
            f'1. Priorizar las URLs con más tráfico GSC de este listado. '
            '2. Medir TTFB real en PageSpeed Insights y WebPageTest. '
            '3. Activar/corregir caché de página (full-page caching). '
            '4. Revisar queries lentas con herramientas de profiling (Query Monitor, New Relic). '
            '5. Evaluar upgrade de hosting o activar CDN con edge caching.',
            f'SF > All URLs > Columna Response Time > ordenar desc. '
            'Filtrar Content Type = text/html + Status = 200 + Indexability = Indexable.',
            'Bajo', 'Alto', 'Medio', 'Dev + Hosting',
            f'Re-crawl: 0 URLs indexables con response_time >{T_SLOW_RESPONSE}s. '
            'PageSpeed Insights: TTFB <800ms en las URLs objetivo.',
            sample_urls(df_slow.sort_values('response_time', ascending=False))
        )

    # T24 — URLs con parámetros indexables (P2)
    n_parametered = len(df_parametered)
    if n_parametered > 0:
        add_task(
            'T24', 'P2', 'URL / Parámetros indexables',
            f'Revisar {n_parametered:,} URLs con parámetros (?) indexadas por Google',
            f'{n_parametered:,} páginas indexables contienen "?" en la URL. '
            'Pueden generar contenido duplicado y desperdiciar crawl budget.',
            f'{n_parametered:,} URLs indexables con al menos un parámetro en la query string. '
            'Google puede tratar cada combinación de parámetros como una URL única, '
            'multiplicando el contenido duplicado y diluyendo la autoridad de las páginas canónicas.',
            'Filtros de tienda/catálogo sin canonical o noindex. '
            'Tracking UTMs en URLs indexables. '
            'Sistema de sesiones o caches con parámetros no filtrados.',
            '1. Clasificar los parámetros: tracking (UTM→ GSC > Parámetros de URL o canonical). '
            '2. Para parámetros de filtro/orden: añadir canonical a la URL sin parámetros. '
            '3. Para parámetros que cambian contenido: evaluar noindex o consolidar con pagination. '
            '4. Configurar GSC > Parámetros de URL para indicar a Google cuáles ignorar.',
            'SF > All URLs > filtrar URL contains "?". Confirmar Indexability = Indexable.',
            'Bajo', 'Medio', 'Bajo', 'SEO + Dev',
            'Re-crawl: las URLs con parámetros tienen canonical correcta o noindex. '
            'GSC Cobertura: reducción de URLs indexadas con parámetros.',
            sample_urls(df_parametered, sort_by='impressions' if HAS_GSC else 'url')
        )

    # T25 — URLs largas >115 caracteres (P2)
    n_long_urls = len(df_long_urls)
    if n_long_urls > 0:
        max_url_len = df_long_urls['url'].str.len().max()
        add_task(
            'T25', 'P2', 'URL / Estructura',
            f'Acortar {n_long_urls:,} URLs indexables con más de {T_URL_MAX_LEN} caracteres',
            f'{n_long_urls:,} URLs indexables superan los {T_URL_MAX_LEN} caracteres. '
            f'URL más larga: {max_url_len} caracteres. '
            'URLs largas tienen menor CTR y son más difíciles de compartir.',
            f'{n_long_urls:,} URLs con len(url) > {T_URL_MAX_LEN} chars en SF. '
            f'La URL más larga tiene {max_url_len} caracteres. '
            'Google muestra como máximo ~70 chars del slug en el snippet; '
            'las URLs largas se truncan y reducen la legibilidad del resultado.',
            'Slugs generados automáticamente desde títulos largos. '
            'Jerarquías de categorías muy anidadas (/cat1/cat2/cat3/producto). '
            'Parámetros incluidos en el slug.',
            '1. Identificar patrones de URL largas (productos, posts, categorías anidadas). '
            '2. Definir convención de slug: máximo 5 palabras clave, sin stop words. '
            '3. Acortar las URLs más visitadas y añadir 301 desde la URL antigua. '
            '4. Actualizar el CMS para generar slugs cortos por defecto.',
            f'SF > All URLs > añadir columna URL Length > filtrar > {T_URL_MAX_LEN} chars.',
            'Medio', 'Bajo', 'Bajo', 'SEO + CMS',
            f'Re-crawl: porcentaje de URLs >{T_URL_MAX_LEN} chars se reduce en ≥50%. '
            'CTR medio mejora en las URLs acortadas (monitorizar en GSC).',
            sample_urls(df_long_urls.assign(_len=df_long_urls['url'].str.len()).sort_values('_len', ascending=False).drop(columns=['_len']))
        )

    # T26 — Titles duplicados (P1)
    n_dup_titles = len(df_dup_titles)
    if HAS_TITLE_1 and n_dup_titles > 0:
        add_task(
            'T26', 'P1', 'Metadatos / Titles duplicados',
            f'Diversificar {n_dup_titles:,} páginas con title tag duplicado '
            f'({n_unique_dup_titles:,} títulos compartidos por varias URLs)',
            f'{n_dup_titles:,} páginas indexables comparten title tag con al menos otra URL. '
            'Google no puede diferenciar el contenido de cada página y puede mergear señales.',
            f'{n_dup_titles:,} páginas indexables tienen un title_1 idéntico a otra URL del sitio. '
            f'Hay {n_unique_dup_titles:,} títulos distintos afectados. '
            'Los titles duplicados dificultan que Google entienda qué página rankear para cada query; '
            'puede elegir la "equivocada" o reescribir el título automáticamente.',
            'Plantillas de CMS generando el mismo title para múltiples páginas. '
            'Páginas de paginación con el mismo title que la página /1. '
            'Productos de la misma categoría sin diferenciador en el title.',
            '1. Exportar el listado y agrupar por title duplicado para identificar patrones. '
            '2. Para páginas de listing/categoría: añadir el número de página al title (ej. "- Página 2"). '
            '3. Para productos: incluir SKU, talla o color. '
            '4. Actualizar la plantilla del CMS para generar títulos únicos automáticamente.',
            'SF > Page Titles > Duplicate + Missing. Filtrar Indexability = Indexable.',
            'Bajo', 'Alto', 'Bajo', 'SEO + CMS',
            'Re-crawl: 0 titles duplicados en páginas indexables. '
            'GSC: mejora en CTR de las páginas con título diferenciado.',
            sample_urls(df_dup_titles, sort_by='impressions' if HAS_GSC else 'url')
        )

    # T27 — Páginas noindex con impresiones en GSC (P1, requiere GSC)
    n_noindex_gsc = len(df_noindex_with_gsc)
    if HAS_GSC and n_noindex_gsc > 0:
        total_impr_noindex = int(df_noindex_with_gsc['impressions'].sum())
        top_noindex_url   = df_noindex_with_gsc.nlargest(1, 'impressions').iloc[0]['url']
        top_noindex_impr  = int(df_noindex_with_gsc['impressions'].max())
        add_task(
            'T27', 'P1', 'Indexación / GSC vs noindex',
            f'Revisar {n_noindex_gsc:,} páginas bloqueadas a la indexación que tienen '
            f'impresiones en GSC ({total_impr_noindex:,} impresiones en total)',
            f'{n_noindex_gsc:,} URLs están marcadas como no-indexables en SF pero '
            f'aún generan {total_impr_noindex:,} impresiones en Google. '
            'Señal de conflicto entre directivas on-page y el índice real de Google.',
            f'{n_noindex_gsc:,} URLs con indexability ≠ Indexable y impressions GSC > 0. '
            f'Total impresiones perdidas: {total_impr_noindex:,}. '
            f'URL con más impresiones: {top_noindex_url} ({top_noindex_impr:,} impr.). '
            'Esto indica que Google tenía estas URLs indexadas antes de añadir el noindex, '
            'o que el noindex se ha aplicado por error.',
            'Noindex añadido por error (ej. staging copiado a producción). '
            'Cambio de estrategia de indexación sin revisar el impacto en tráfico activo. '
            'Canonical a otra URL en páginas que sí deben indexar.',
            '1. Para cada URL: decidir si el noindex es intencionado o un error. '
            '2. Si es error: eliminar el noindex/canonical y monitorizar re-indexación en GSC. '
            '3. Si es intencionado: redirigir el tráfico residual a una URL equivalente indexable. '
            '4. Revisar el proceso de QA para evitar noindex en producción.',
            'SF > Indexability > Non-Indexable. Cruzar con GSC > Rendimiento. '
            'Filtrar páginas con impressions > 0.',
            'Bajo', 'Alto', 'Bajo', 'SEO',
            'Re-crawl: las URLs con tráfico GSC son Indexable. '
            'GSC Cobertura: reducción de "Excluida - noindex" en URLs con impresiones.',
            sample_urls(df_noindex_with_gsc.sort_values('impressions', ascending=False))
        )

    # T28 — Titles largos >60 chars (P2)
    n_long_titles = len(df_long_titles)
    if HAS_TITLE_LEN and n_long_titles > 0:
        avg_title_len = round(df_long_titles['title_1_length'].mean(), 0) if 'title_1_length' in df_long_titles.columns else 0
        add_task(
            'T28', 'P2', 'Metadatos / Title largo',
            f'Acortar title tag en {n_long_titles:,} páginas indexables (>{60} caracteres)',
            f'{n_long_titles:,} páginas tienen un title de más de 60 caracteres '
            f'(media: {avg_title_len:.0f} chars). Google trunca a ~60 chars y puede reescribir el título.',
            f'{n_long_titles:,} URLs indexables con title_1_length > 60. '
            f'Media de longitud en las afectadas: {avg_title_len:.0f} caracteres. '
            'Google trunca el title en el snippet a ~600 px (≈60 chars); '
            'un title truncado pierde la keyword final y puede bajar el CTR.',
            'Plantillas de CMS usando nombre completo del producto/post sin límite de caracteres. '
            'Textos de title generados manualmente sin revisión de longitud. '
            'Titles con boilerplate al final (ej. "| Nombre de la tienda | Categoría").',
            '1. Exportar el listado y priorizar las páginas con más impresiones. '
            '2. Reescribir el title para que la keyword principal aparezca en los primeros 60 chars. '
            '3. Eliminar boilerplate superfluo o moverlo al inicio si es relevante. '
            '4. Configurar la plantilla del CMS para avisar cuando el title supere 60 chars.',
            'SF > Page Titles > Over 60 Characters (o columna Title 1 Length > 60).',
            'Bajo', 'Medio', 'Bajo', 'SEO',
            'Re-crawl: todas las páginas indexables tienen title_1_length ≤ 60. '
            'GSC: mejora de CTR en las URLs corregidas.',
            sample_urls(df_long_titles.sort_values('title_1_length', ascending=False) if 'title_1_length' in df_long_titles.columns else df_long_titles)
        )

    # T29 — Canonical ausente en páginas 200 indexables (P2)
    n_no_canonical = len(df_no_canonical)
    if HAS_CANONICAL_COL and n_no_canonical > 0:
        add_task(
            'T29', 'P2', 'Canonicals / Ausente',
            f'Añadir canonical self-referencing en {n_no_canonical:,} páginas indexables sin canonical',
            f'{n_no_canonical:,} páginas con status 200 e indexabilidad "Indexable" '
            'no tienen etiqueta canonical. Sin canonical, URLs con parámetros o variantes '
            'pueden generar duplicados no controlados.',
            f'{n_no_canonical:,} URLs con status 200, Indexable y canonical_link_element_1 vacío en SF. '
            'La ausencia de canonical self-referencing permite que variaciones de la URL '
            '(con parámetros de tracking, sesión, UTM…) sean tratadas como URLs independientes '
            'por Googlebot, diluyendo señales de PageRank.',
            'CMS sin implementación de canonical por defecto. '
            'Páginas creadas manualmente fuera del flujo habitual del CMS. '
            'Actualizaciones del CMS que borraron la lógica de canonical.',
            '1. Verificar si el CMS implementa canonical de forma global. '
            '2. Si no: añadir <link rel="canonical" href="{url_actual}"> en el <head> de todas las páginas. '
            '3. Para WordPress/Shopify: activar canonical en el plugin SEO (Yoast, Rank Math, etc.). '
            '4. Re-crawl para confirmar que canonical_link_element_1 = url en todas las páginas.',
            'SF > Canonicals > Missing. Filtrar Status 200 + Indexability = Indexable.',
            'Bajo', 'Medio', 'Bajo', 'Dev + SEO',
            'Re-crawl: 0 páginas indexables sin canonical. '
            'GSC Cobertura: reducción de variantes de URL duplicadas.',
            sample_urls(df_no_canonical, sort_by='impressions' if HAS_GSC else 'url')
        )

    # T30 — Sin H2 en páginas SEO indexables (P2)
    n_no_h2 = len(df_no_h2)
    if HAS_H2_COL and n_no_h2 > 0:
        add_task(
            'T30', 'P2', 'Contenido / Estructura H2',
            f'Añadir H2 a {n_no_h2:,} páginas SEO indexables sin subtítulos',
            f'{n_no_h2:,} páginas de tipo SEO (colecciones, productos, posts) '
            'carecen de etiqueta H2. La ausencia de subtítulos dificulta la comprensión '
            'del contenido por parte de Google.',
            f'{n_no_h2:,} URLs de tipo SEO con h2_1 vacío en SF. '
            'Google usa los H2 para entender la estructura temática de la página '
            'y para generar featured snippets. Sin H2, el contenido es menos escaneable '
            'tanto para usuarios como para bots.',
            'Plantillas de CMS que no incluyen H2 por defecto. '
            'Contenido migrado sin revisar la estructura de headings. '
            'Páginas con contenido mínimo en las que no se añadieron subtítulos.',
            '1. Identificar qué tipos de página afectan más (productos, posts, categorías). '
            '2. Actualizar las plantillas para incluir al menos un H2 con keyword secundaria. '
            '3. Para páginas de alto tráfico: revisar manualmente y redactar subtítulos relevantes. '
            '4. Verificar en re-crawl que h2_1 ya no está vacío.',
            'SF > Headings > H2 Missing or Empty. Filtrar url_type = SEO_TYPES.',
            'Bajo', 'Medio', 'Bajo', 'SEO + CMS',
            'Re-crawl: 0 páginas SEO sin H2. '
            'GSC: mejora en featured snippets y posición media en las URLs corregidas.',
            sample_urls(df_no_h2, sort_by='impressions' if HAS_GSC else 'url')
        )

    # T31 — Title = H1 exacto (P2)
    n_title_eq_h1 = len(df_title_eq_h1)
    if HAS_TITLE_1 and HAS_H1_COL and n_title_eq_h1 > 0:
        add_task(
            'T31', 'P2', 'Metadatos / Title = H1 duplicado',
            f'Diferenciar title tag y H1 en {n_title_eq_h1:,} páginas con texto idéntico',
            f'{n_title_eq_h1:,} páginas tienen el title tag y el H1 con exactamente el mismo texto. '
            'Es una oportunidad perdida: el title puede optimizarse para CTR '
            'y el H1 para la relevancia temática en página.',
            f'{n_title_eq_h1:,} URLs donde title_1 == h1_1 (comparación sin case, sin spaces). '
            'El title se muestra en el SERP y debe maximizar el CTR; el H1 es el heading '
            'principal de la página y debe contextualizar el contenido para el usuario. '
            'Usarlos idénticos no es un error técnico, pero limita la optimización semántica.',
            'Plantillas de CMS que auto-generan el H1 desde el title o viceversa. '
            'Falta de revisión editorial al crear nuevas páginas.',
            '1. Para cada página: redactar el title orientado al CTR en SERP '
            '(incluir número, año, benefit o diferenciador). '
            '2. Mantener el H1 más descriptivo y conversacional para el usuario en página. '
            '3. Ambos deben contener la keyword principal, pero con variación de redacción.',
            'SF > Page Titles > comparar con H1 Headings manualmente o exportar y cruzar en Excel.',
            'Bajo', 'Bajo', 'Bajo', 'SEO',
            'Re-crawl: las páginas corregidas tienen title ≠ H1. '
            'GSC: mejora de CTR en las URLs modificadas.',
            sample_urls(df_title_eq_h1, sort_by='impressions' if HAS_GSC else 'url')
        )

    # T32 — Crawl depth alto >4 con tráfico GSC (P2)
    n_deep = len(df_deep)
    if HAS_DEPTH_COL and n_deep > 0:
        avg_depth = round(df_deep['depth'].mean(), 1) if 'depth' in df_deep.columns else 0
        add_task(
            'T32', 'P2', 'Arquitectura / Crawl Depth',
            f'Reducir profundidad de rastreo en {n_deep:,} URLs importantes (>{T_MAX_DEPTH} clicks)',
            f'{n_deep:,} páginas indexables con señal SEO están a más de {T_MAX_DEPTH} clicks '
            f'de la home (media: {avg_depth} clicks). Googlebot les asigna menor prioridad de rastreo.',
            f'{n_deep:,} URLs con depth > {T_MAX_DEPTH} y señal SEO (impressions > 0 o url_type = SEO_TYPES). '
            f'Profundidad media de las afectadas: {avg_depth} clicks. '
            'Las páginas a más de 4 clicks de la home reciben menos PageRank '
            'y menos frecuencia de rastreo. Regla general: ninguna página importante '
            'debería estar a más de 3-4 clicks.',
            'Arquitectura de categorías muy anidada. '
            'Paginaciones profundas o filtros generando rutas largas. '
            'Enlazado interno insuficiente que no acorta el camino desde la home.',
            '1. Exportar el listado y priorizar por impresiones GSC. '
            '2. Para las URLs más importantes: añadir enlace directo desde la home, '
            'categorías principales o el menú de navegación. '
            '3. Revisar si la estructura de carpetas/categorías puede aplanarse. '
            '4. Añadir las URLs más profundas al sitemap (ya incluidas si T22 se ejecutó).',
            f'SF > All URLs > columna Crawl Depth > filtrar > {T_MAX_DEPTH}. '
            'Cruzar con GSC Rendimiento para priorizar las más vistas.',
            'Bajo', 'Medio', 'Medio', 'SEO + Dev',
            f'Re-crawl: reducción ≥30% de URLs con depth > {T_MAX_DEPTH}. '
            'GSC: mejora de frecuencia de rastreo en las URLs corregidas.',
            sample_urls(df_deep.sort_values('impressions', ascending=False) if HAS_GSC and 'impressions' in df_deep.columns else df_deep)
        )

    print(f"  Total tareas: {len(tasks)}")


    # ──────────────────────────────────────────────────────────────────────────────
    # 7. URLs-PRIORIDAD
    # ──────────────────────────────────────────────────────────────────────────────
    print("\nBuilding URLs-Prioridad...")

    url_prio_rows = []

    def add_url_prio(df_sub, issue_label, prioridad, action, max_n=50, sort_by='impressions'):
        if len(df_sub) == 0:
            return
        df_sub = df_sub.copy()
        if sort_by in df_sub.columns:
            df_sub = df_sub.sort_values(sort_by, ascending=False)
        for _, r in df_sub.head(max_n).iterrows():
            s   = int(r.get('status', 0))
            idx = str(r.get('indexable', r.get('indexability_status', 'N/D')))[:20]
            il  = int(r.get('inlinks', 0))
            d   = int(r.get('depth', 0))
            metrica_sf = f"Status {s} | {idx} | Inlinks {il} | Depth {d}"

            impr    = int(r.get('impressions', 0))
            clicks  = int(r.get('clicks', 0))
            ctr_val = r.get('ctr', 0)
            ctr_pct = ctr_val * 100 if ctr_val <= 1 else ctr_val
            pos     = round(float(r.get('position', 0)), 2)
            gsc_str = f"Impr {impr:,} | Clics {clicks} | CTR {ctr_pct:.1f}% | Pos {pos}" if impr > 0 else "Sin datos GSC"

            url_prio_rows.append({
                'URL': r['url'],
                'Issue principal': issue_label,
                'Prioridad': prioridad,
                'Métricas SF': metrica_sf,
                'Señal GSC': gsc_str,
                'Acción recomendada': action,
            })


    # P0: locale especial
    if SPECIAL_LOCALE and len(df_ca) > 0:
        _sort_ca = 'impressions' if HAS_GSC and 'impressions' in df_ca.columns else None
        add_url_prio(
            df_ca.sort_values(_sort_ca, ascending=False) if _sort_ca else df_ca,
            f'Locale {SPECIAL_LOCALE.upper()} detectado (resto bloqueado por robots.txt)', 'P0',
            f'Confirmar si /{SPECIAL_LOCALE}/ debe indexar. '
            f'Si sí: eliminar Disallow: /{SPECIAL_LOCALE}/ de robots.txt.',
            max_n=5
        )

    # P0: 503
    if n_503 > 0:
        _sort503 = 'inlinks' if HAS_INLINKS_DATA else 'impressions'
        add_url_prio(df_503.sort_values(_sort503, ascending=False) if _sort503 in df_503.columns else df_503,
                     'Error 503 durante crawl', 'P0',
                     'Verificar disponibilidad en producción. Si persiste, escalar a Dev.', max_n=15)

    # P1: 301
    _sort301 = 'inlinks' if HAS_INLINKS_DATA else ('impressions' if HAS_GSC else None)
    add_url_prio(df_301.sort_values(_sort301, ascending=False) if _sort301 and _sort301 in df_301.columns else df_301,
                 'URL redirigida (301) — actualizar enlaces internos', 'P1',
                 'Actualizar enlace(s) para apuntar a la URL destino final.', max_n=30)

    # P1: 404
    _sort404 = 'inlinks' if HAS_INLINKS_DATA else ('impressions' if HAS_GSC else None)
    add_url_prio(df_404.sort_values(_sort404, ascending=False) if _sort404 and _sort404 in df_404.columns else df_404,
                 'URL con error 404 detectada en crawl', 'P1',
                 'Crear redirect 301 a URL equivalente. SF > All Inlinks para priorizar.', max_n=50)

    # P1: double slash (todas las HTML)
    add_url_prio(df_double_slash_all,
                 'URL con doble slash (//) — posible duplicado', 'P1',
                 'Añadir redirect 301 de URL con // a la versión sin //. Corregir en template.', max_n=50, sort_by='inlinks')

    # P1: sin meta desc con impresiones
    if HAS_GSC:
        df_no_meta_impr = df_no_meta[df_no_meta['impressions'] > 0].sort_values('impressions', ascending=False)
        add_url_prio(df_no_meta_impr, 'Sin meta description — URL con tráfico orgánico', 'P1',
                     'Redactar meta description única (120-155 chars) con keyword + CTA.', max_n=30)

    # P1: sin H1 con impresiones
    if HAS_GSC:
        df_no_h1_impr = df_no_h1[df_no_h1['impressions'] > 0].sort_values('impressions', ascending=False)
        add_url_prio(df_no_h1_impr, 'Sin H1 — URL relevante para SEO', 'P1',
                     'Añadir <h1> con keyword principal como primer heading de la página.', max_n=20)

    # P1: cart-reco indexables
    if HAS_CART_RECO and len(df_cart_reco_indexable) > 0:
        add_url_prio(df_cart_reco_indexable,
                     'Cart-recommendations indexable — no debe aparecer en Google', 'P1',
                     'Añadir meta robots noindex o bloquear en robots.txt.', max_n=10)

    # P1: paginations indexables con impresiones
    if HAS_GSC:
        df_pag_gsc = df_paginaciones_indexable[df_paginaciones_indexable['impressions'] > 0]
        if len(df_pag_gsc) > 0:
            add_url_prio(df_pag_gsc, 'Paginación indexada con señal GSC — candidata a canonical', 'P1',
                         'Añadir canonical a la página /1. Evaluar noindex en page/2+.', max_n=20)

    # P1: pos 11-20
    if len(df_pos_11_20) > 0:
        add_url_prio(df_pos_11_20.sort_values('impressions', ascending=False),
                     'Posición 11-20: candidata a página 1', 'P1',
                     'Mejorar contenido, reforzar enlazado interno y optimizar snippet para Top 10.', max_n=30)

    # P1: CTR < umbral en top 10
    if len(df_low_ctr) > 0:
        add_url_prio(df_low_ctr.sort_values('impressions', ascending=False),
                     f'Top 10 con CTR <{T_CTR_LOW*100:.0f}% — snippet poco atractivo', 'P1',
                     'Reescribir title con diferenciador. Meta description con CTA. Evaluar rich snippets.', max_n=25)

    # P1: noindex en sitemap
    if HAS_SITEMAP_DATA and len(df_noindex_in_sitemap) > 0:
        _sort_ni = 'impressions' if HAS_GSC and 'impressions' in df_noindex_in_sitemap.columns else 'inlinks'
        _df_ni = df_noindex_in_sitemap.sort_values(_sort_ni, ascending=False) if _sort_ni in df_noindex_in_sitemap.columns else df_noindex_in_sitemap
        add_url_prio(_df_ni, 'URL no-indexable incluida en sitemap XML', 'P1',
                     'Eliminar URL del sitemap.xml o corregir la directiva noindex/canonical.', max_n=30)

    # P1: noindex con impresiones GSC
    if HAS_GSC and len(df_noindex_with_gsc) > 0:
        add_url_prio(df_noindex_with_gsc.sort_values('impressions', ascending=False),
                     'Página noindex con impresiones orgánicas — posible pérdida de tráfico', 'P1',
                     'Revisar si el noindex es intencional. Si debe indexar: eliminar directiva y añadir al sitemap.', max_n=25)

    # P1: páginas lentas >3s
    if HAS_RESPONSE_TIME and len(df_slow) > 0:
        _sort_slow = 'impressions' if HAS_GSC and 'impressions' in df_slow.columns else 'inlinks'
        _df_slow_s = df_slow.sort_values(_sort_slow, ascending=False) if _sort_slow in df_slow.columns else df_slow
        add_url_prio(_df_slow_s,
                     f'Tiempo de respuesta >{T_SLOW_RESPONSE:.0f}s — penalización Core Web Vitals', 'P1',
                     'Optimizar TTFB: caché servidor, CDN, reducir queries. Objetivo <1.8s.', max_n=20)

    # P1: titles duplicados
    if HAS_TITLE_1 and len(df_dup_titles) > 0:
        _sort_dt = 'impressions' if HAS_GSC and 'impressions' in df_dup_titles.columns else 'inlinks'
        _df_dt = df_dup_titles.sort_values(_sort_dt, ascending=False) if _sort_dt in df_dup_titles.columns else df_dup_titles
        add_url_prio(_df_dt, 'Title duplicado — Google puede ignorar uno de los documentos', 'P1',
                     'Redactar title único por URL con keyword principal diferenciada.', max_n=20)

    # P2: crawl depth alto con tráfico
    if HAS_DEPTH_COL and len(df_deep) > 0:
        _sort_deep = 'impressions' if HAS_GSC and 'impressions' in df_deep.columns else 'inlinks'
        _df_deep_s = df_deep.sort_values(_sort_deep, ascending=False) if _sort_deep in df_deep.columns else df_deep
        add_url_prio(_df_deep_s,
                     f'Crawl depth >{T_MAX_DEPTH} — página importante pero difícil de rastrear', 'P2',
                     'Añadir enlaces internos desde páginas de mayor jerarquía. Revisar arquitectura de categorías.', max_n=20)

    # P2: URLs indexables ausentes del sitemap (muestra las de mayor tráfico)
    if HAS_SITEMAP_DATA and len(df_missing_from_sitemap) > 0:
        _sort_ms = 'impressions' if HAS_GSC and 'impressions' in df_missing_from_sitemap.columns else 'inlinks'
        _df_ms = df_missing_from_sitemap.sort_values(_sort_ms, ascending=False) if _sort_ms in df_missing_from_sitemap.columns else df_missing_from_sitemap
        add_url_prio(_df_ms, 'URL indexable ausente del sitemap XML', 'P2',
                     'Añadir URL al sitemap.xml para facilitar descubrimiento por Googlebot.', max_n=20)

    # Deduplicar por URL, conservar la de mayor prioridad
    prio_order = {'P0': 0, 'P1': 1, 'P2': 2, 'P3': 3}
    seen = {}
    for row in url_prio_rows:
        url = row['URL']
        if url not in seen or prio_order.get(row['Prioridad'], 9) < prio_order.get(seen[url]['Prioridad'], 9):
            seen[url] = row

    def _sort_key(r):
        p = prio_order.get(r['Prioridad'], 9)
        if 'Impr' in r['Señal GSC'] and r['Señal GSC'] != 'Sin datos GSC':
            try:
                impr = int(r['Señal GSC'].split('Impr ')[1].split(' |')[0].replace(',', ''))
            except (IndexError, ValueError):
                impr = 0
        else:
            impr = 0
        return (p, -impr)

    url_prio_final = sorted(seen.values(), key=_sort_key)
    print(f"  URLs-Prioridad: {len(url_prio_final)} URLs únicas")


    # ──────────────────────────────────────────────────────────────────────────────
    # 8. OPORTUNIDADES GSC
    # ──────────────────────────────────────────────────────────────────────────────
    print("\nBuilding Oportunidades GSC...")

    gsc_rows = []

    def add_gsc_opp(df_sub, tipo_opp, accion, max_n=80):
        if len(df_sub) == 0:
            return
        for _, r in df_sub.nlargest(max_n, 'impressions').iterrows():
            ctr_val = r.get('ctr', 0)
            ctr_pct = ctr_val * 100 if ctr_val <= 1 else ctr_val
            gsc_rows.append({
                'URL': r['url'],
                'Query': 'N/D (solo GSC por URL en SF)',
                'Impresiones': int(r['impressions']),
                'Clics': int(r['clicks']),
                'CTR': round(ctr_pct / 100, 4),
                'Posición': round(float(r.get('position', 0)), 2),
                'Tipo oportunidad': tipo_opp,
                'Acción recomendada': accion,
            })


    if HAS_GSC:
        add_gsc_opp(df_pos_11_20, 'Posición 11-20',
                    'Optimizar contenido/snippet y reforzar enlazado interno para Top 10.', max_n=80)
        add_gsc_opp(df_low_ctr, f'CTR <{T_CTR_LOW*100:.0f}% en Top 10',
                    'Reescribir title con diferenciador. Meta description con CTA. Considerar schema.', max_n=50)
        add_gsc_opp(df_impr_no_clicks, f'≥{T_IMPRESSIONS_NO_CLICKS} impresiones, 0 clics',
                    'Revisar intent: ¿responde la página a lo que busca el usuario? Mejorar snippet urgente.', max_n=30)

    gsc_seen = {}
    for row in gsc_rows:
        if row['URL'] not in gsc_seen:
            gsc_seen[row['URL']] = row

    gsc_final = sorted(gsc_seen.values(), key=lambda r: -r['Impresiones'])
    print(f"  Oportunidades GSC: {len(gsc_final)} URLs")


    # ──────────────────────────────────────────────────────────────────────────────
    # 9. RESUMEN
    # ──────────────────────────────────────────────────────────────────────────────
    print("\nBuilding Resumen...")

    def r(label, value):
        return {'Métrica': label, 'Valor': value}

    resumen_rows = []

    # Inventario
    resumen_rows += [
        r('▌ INVENTARIO', ''),
        r('Archivo procesado', RUTA_CSV.split('/')[-1]),
        r('Fecha de análisis', datetime.now().strftime('%Y-%m-%d')),
        r('Total filas en CSV', f"{len(df_raw):,}"),
        r('HTML filtradas (text/html)', f"{len(df):,}"),
        r('Datasets disponibles', 'SF + GSC integrado' if HAS_GSC else 'SF (sin datos GSC)'),
        r('', ''),
    ]

    # Cobertura del crawl
    resumen_rows += [
        r('▌ COBERTURA DEL CRAWL', ''),
        r('Total URLs HTML en crawl', f"{total_html:,}"),
        r('Status 200 (OK)', f"{status_200:,}"),
        r('Status 0 (sin respuesta)', f"{status_0:,}"),
        r('Status 301 (redirigidas)', f"{status_301_count:,}"),
        r('Status 404 (no encontradas)', f"{status_404_count:,}"),
        r('Status 5xx (errores servidor)', f"{status_5xx_count:,}"),
        r('', ''),
    ]

    # Indexación
    resumen_rows += [
        r('▌ INDEXACIÓN', ''),
        r('URLs indexables', f"{total_indexable:,}"),
        r('URLs no indexables', f"{total_no_indexable:,}"),
        r('  — Bloqueadas por robots.txt', f"{total_blocked_robots:,}"),
        r('  — Canonicalizadas', f"{total_canonicalized:,}"),
        r('  — Redirigidas', f"{total_redirects_idx:,}"),
        r('  — Con doble slash (//) detectadas', f"{n_ds:,}"),
        r('  — Paginaciones detectadas', f"{n_pag_total:,}"),
        r('  — Paginaciones indexadas', f"{n_pag_index:,}"),
    ]
    if HAS_CART_RECO:
        resumen_rows.append(r('  — Cart-recommendations indexadas', f"{len(df_cart_reco_indexable):,}"))
    resumen_rows.append(r('', ''))

    # Internacional (solo si multilingual)
    if IS_MULTILINGUAL and SITE_LANGS:
        lang_breakdown = ', '.join([f"/{l}/: {lang_indexable.get(l, 0):,}" for l in SITE_LANGS])
        resumen_rows += [
            r('▌ INTERNACIONAL', ''),
            r('Idiomas detectados en crawl', ', '.join(['raíz'] + [f'/{l}/' for l in SITE_LANGS])),
            r('Idiomas con páginas indexables', ', '.join(['root'] + langs_with_indexable)),
            r('Indexables por idioma', f"raíz: {lang_indexable.get('root', 0):,} | {lang_breakdown}"),
        ]
        if SPECIAL_LOCALE:
            resumen_rows += [
                r(f'/{SPECIAL_LOCALE}/ URLs en crawl', f"{ca_url_count:,}"),
                r(f'/{SPECIAL_LOCALE}/ bloqueadas robots.txt', f"N/D — robots.txt bloquea el crawl"),
                r(f'Nota /{SPECIAL_LOCALE}/', f'Ejecutar SF con robots.txt OFF para descubrir todas las URLs /{SPECIAL_LOCALE}/'),
                r('Decisión pendiente', f'Si /{SPECIAL_LOCALE}/ debe indexarse: eliminar Disallow (P0, ver T01)'),
            ]
        resumen_rows.append(r('', ''))

    # Calidad on-page
    resumen_rows += [
        r('▌ CALIDAD ON-PAGE', ''),
        r('Páginas SEO sin meta description', f"{n_no_meta:,}"),
        r('  — Productos sin meta desc', f"{n_no_meta_products:,}"),
        r('  — Colecciones/categorías sin meta desc', f"{n_no_meta_collections:,}"),
        r('  — Páginas estáticas sin meta desc', f"{n_no_meta_pages:,}"),
        r('Páginas SEO sin H1', f"{n_no_h1:,}"),
        r('  — Páginas/posts sin H1', f"{n_no_h1_info:,}"),
        r('Titles cortos (<{} chars) en colecciones'.format(T_TITLE_SHORT_CHARS), f"{n_short_title:,}"),
        r('301 con inlinks (requiere All Inlinks export)', f"{'N/D — exportar SF > All Inlinks' if not HAS_INLINKS_DATA else str(len(df_301[df_301['inlinks'] > 0])) + ' detectadas'}"),
        r('301 total en crawl', f"{status_301_count:,}"),
        r('404 total en crawl', f"{status_404_count:,}"),
        r('', ''),
    ]

    # Señales GSC
    if HAS_GSC:
        resumen_rows += [
            r('▌ SEÑALES GSC', ''),
            r('URLs con datos GSC', f"{len(df_gsc):,}"),
            r('Total impresiones (periodo)', f"{total_impressions:,}"),
            r('Total clics (periodo)', f"{total_clicks:,}"),
            r('CTR global', f"{overall_ctr:.2%}"),
            r(f'URLs en posición {T_POS_OPP_MIN}-{T_POS_OPP_MAX}', f"{len(df_pos_11_20):,}"),
            r(f'URLs top 10 con CTR <{T_CTR_LOW*100:.0f}%', f"{len(df_low_ctr):,}"),
            r(f'URLs con ≥{T_IMPRESSIONS_NO_CLICKS} impr y 0 clics', f"{len(df_impr_no_clicks):,}"),
            r('URLs con demanda y bajo enlazado', f"{n_demand:,}"),
            r('', ''),
        ]

    # Técnico avanzado T21-T32
    resumen_rows.append(r('▌ TÉCNICO AVANZADO (T21–T32)', ''))
    # Sitemaps
    if HAS_SITEMAP_DATA:
        resumen_rows.append(r('URLs noindex incluidas en sitemap', f"{len(df_noindex_in_sitemap):,}"))
        resumen_rows.append(r('URLs indexables ausentes del sitemap', f"{len(df_missing_from_sitemap):,}"))
    else:
        resumen_rows.append(r('Datos de sitemap', 'N/D — activar "Crawl linked XML Sitemaps" en SF'))
    # Velocidad
    if HAS_RESPONSE_TIME:
        resumen_rows.append(r(f'Páginas lentas >{T_SLOW_RESPONSE:.0f}s (status 200 indexable)', f"{len(df_slow):,}"))
    else:
        resumen_rows.append(r('Tiempo de respuesta', 'N/D — activar "Response Time" en SF'))
    # URLs
    resumen_rows.append(r(f'URLs indexables con parámetros (?)', f"{len(df_parametered):,}"))
    resumen_rows.append(r(f'URLs indexables largas (>{T_URL_MAX_LEN} chars)', f"{len(df_long_urls):,}"))
    # Metadatos
    if HAS_TITLE_1:
        resumen_rows.append(r('Páginas con title duplicado', f"{len(df_dup_titles):,} URLs ({n_unique_dup_titles} títulos únicos duplicados)"))
    if HAS_TITLE_LEN:
        resumen_rows.append(r('Titles largos (>60 chars) en indexables', f"{len(df_long_titles):,}"))
    if HAS_TITLE_1 and HAS_H1_COL:
        resumen_rows.append(r('Title = H1 exacto (páginas indexables)', f"{len(df_title_eq_h1):,}"))
    # GSC cruzado
    if HAS_GSC:
        resumen_rows.append(r('Páginas noindex con impresiones GSC', f"{len(df_noindex_with_gsc):,}"))
    # Canonicals / estructura
    if HAS_CANONICAL_COL:
        resumen_rows.append(r('Páginas 200 indexables sin canonical', f"{len(df_no_canonical):,}"))
    # Contenido
    if HAS_H2_COL:
        resumen_rows.append(r('Páginas SEO sin H2', f"{len(df_no_h2):,}"))
    # Arquitectura
    if HAS_DEPTH_COL:
        resumen_rows.append(r(f'Páginas indexables con depth >{T_MAX_DEPTH}', f"{len(df_deep):,}"))
    resumen_rows.append(r('', ''))

    # Top P0
    p0_tasks = [t for t in tasks if t['Prioridad'] == 'P0']
    resumen_rows.append(r('▌ TOP P0 — CRÍTICO (acción inmediata)', ''))
    if p0_tasks:
        for i, t in enumerate(p0_tasks, 1):
            resumen_rows.append(r(f'P0-{i}', f"[{t['ID']}] {t['Tarea']} — {t['Evidencia'][:200]}"))
    else:
        resumen_rows.append(r('P0', 'Sin issues P0 detectados'))
    resumen_rows.append(r('', ''))

    # Top P1
    p1_tasks = [t for t in tasks if t['Prioridad'] == 'P1']
    resumen_rows.append(r('▌ TOP P1 — IMPORTANTE (próximas semanas)', ''))
    for i, t in enumerate(p1_tasks[:6], 1):
        resumen_rows.append(r(f'P1-{i}', f"[{t['ID']}] {t['Tarea']}"))
    resumen_rows.append(r('', ''))

    # Pendientes
    resumen_rows += [
        r('▌ PENDIENTES DE CONFIRMAR (requieren datos adicionales)', ''),
        r('Sitemaps XML', 'No incluido en el export de SF. Verificar en /sitemap.xml.'),
        r('Hreflang', 'Export SF no incluye columna hreflang. Requiere recrawl con opción específica.' if IS_MULTILINGUAL else 'N/A — IS_MULTILINGUAL=False. Activar si el sitio tiene múltiples idiomas.'),
        r('Schema markup', 'No analizable desde SF exports estándar. Verificar con Rich Results Test.'),
        r('Canibalización', 'No analizable sin queries por URL. Requiere export de GSC Search Analytics o herramienta tipo Semrush/Ahrefs.'),
    ]

    print(f"  Resumen: {len(resumen_rows)} filas")


    # ──────────────────────────────────────────────────────────────────────────────
    # 10. ESCRITURA DEL EXCEL
    # ──────────────────────────────────────────────────────────────────────────────
    print(f"\nGenerating Excel → {OUTPUT_PATH}")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Estilos globales
    header_fill  = PatternFill('solid', fgColor=HEADER_BG)
    header_font  = Font(bold=True, color=WHITE, size=11, name='Calibri')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    body_font    = Font(size=10, name='Calibri')
    wrap_align   = Alignment(vertical='top', wrap_text=True)
    top_align    = Alignment(vertical='top', wrap_text=False)
    thin_side    = Side(style='thin', color='CCCCCC')
    thin_border  = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    PRIO_FILLS = {
        'P0': PatternFill('solid', fgColor=P0_BG),
        'P1': PatternFill('solid', fgColor=P1_BG),
        'P2': PatternFill('solid', fgColor=P2_BG),
        'P3': PatternFill('solid', fgColor=P3_BG),
    }

    def style_header_row(ws, row_num=1, height=28):
        for cell in ws[row_num]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = thin_border
        ws.row_dimensions[row_num].height = height

    def set_col_widths(ws, widths_dict):
        for col_letter, width in widths_dict.items():
            ws.column_dimensions[col_letter].width = width

    def alt_fill(row_idx):
        return PatternFill('solid', fgColor='F0F4F8') if row_idx % 2 == 0 else None


    # ── HOJA 1: RESUMEN ──────────────────────────────────────────────────────────
    ws_res = wb.create_sheet('Resumen')
    ws_res.append(['Métrica', 'Valor'])
    style_header_row(ws_res)

    section_fill = PatternFill('solid', fgColor='2C3E50')
    section_font = Font(bold=True, color=WHITE, size=10, name='Calibri')

    for i, row in enumerate(resumen_rows, start=2):
        ws_res.append([row['Métrica'], row['Valor']])
        cell_m = ws_res.cell(row=i, column=1)
        cell_v = ws_res.cell(row=i, column=2)
        if str(row['Métrica']).startswith('▌'):
            cell_m.fill = section_fill
            cell_m.font = section_font
            cell_v.fill = section_fill
            cell_v.font = section_font
        elif row['Métrica'] != '':
            cell_m.font = Font(size=10, name='Calibri')
            cell_v.font = Font(size=10, name='Calibri')
            if i % 2 == 0:
                cell_m.fill = PatternFill('solid', fgColor='F0F4F8')
                cell_v.fill = PatternFill('solid', fgColor='F0F4F8')
        cell_m.alignment = Alignment(vertical='center', wrap_text=False)
        cell_v.alignment = Alignment(vertical='center', wrap_text=True)

    set_col_widths(ws_res, {'A': 45, 'B': 90})
    ws_res.freeze_panes = 'A2'
    ws_res.sheet_properties.tabColor = '1F4E78'


    # ── HOJA 2: TAREAS ───────────────────────────────────────────────────────────
    ws_tar = wb.create_sheet('Tareas')
    task_headers = ['ID', 'Prioridad', 'Categoría', 'Tarea', 'Descripción corta', 'Evidencia',
                    'Causa probable', 'Qué hacer', 'Dónde detectarlo', 'Esfuerzo', 'Impacto',
                    'Riesgo', 'Responsable', 'Validación', 'URLs ejemplo']
    ws_tar.append(task_headers)
    style_header_row(ws_tar)

    for i, task in enumerate(tasks, start=2):
        row_data = [task[h] for h in task_headers]
        ws_tar.append(row_data)
        prio = task['Prioridad']
        fill = PRIO_FILLS.get(prio)
        for col_idx in range(1, len(task_headers) + 1):
            cell = ws_tar.cell(row=i, column=col_idx)
            if fill:
                cell.fill = fill
            cell.font = body_font
            cell.border = thin_border
            col_letter = get_column_letter(col_idx)
            if col_letter == 'O':
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                cell.font = Font(size=9, name='Calibri', color='1155CC')
            else:
                cell.alignment = wrap_align
        ws_tar.row_dimensions[i].height = 90

    set_col_widths(ws_tar, {
        'A': 8, 'B': 10, 'C': 24, 'D': 28, 'E': 34,
        'F': 42, 'G': 28, 'H': 46, 'I': 28, 'J': 10,
        'K': 10, 'L': 10, 'M': 16, 'N': 28, 'O': 60,
    })
    ws_tar.freeze_panes = 'C2'
    ws_tar.sheet_properties.tabColor = 'C0392B'


    # ── HOJA 3: URLs-PRIORIDAD ───────────────────────────────────────────────────
    ws_url = wb.create_sheet('URLs - Prioridad')
    url_headers = ['URL', 'Issue principal', 'Prioridad', 'Métricas SF', 'Señal GSC', 'Acción recomendada']
    ws_url.append(url_headers)
    style_header_row(ws_url)

    for i, row in enumerate(url_prio_final, start=2):
        ws_url.append([row[h] for h in url_headers])
        prio = row['Prioridad']
        fill = PRIO_FILLS.get(prio)
        for col_idx in range(1, len(url_headers) + 1):
            cell = ws_url.cell(row=i, column=col_idx)
            if fill:
                cell.fill = fill
            cell.font = body_font
            cell.border = thin_border
            col_letter = get_column_letter(col_idx)
            if col_letter == 'A':
                cell.font = Font(size=9, name='Calibri', color='1155CC')
            cell.alignment = top_align
        ws_url.row_dimensions[i].height = 18

    set_col_widths(ws_url, {'A': 70, 'B': 34, 'C': 10, 'D': 36, 'E': 36, 'F': 42})
    ws_url.freeze_panes = 'A2'
    ws_url.sheet_properties.tabColor = 'E67E22'


    # ── HOJA 4: OPORTUNIDADES GSC ────────────────────────────────────────────────
    ws_gsc = wb.create_sheet('Oportunidades GSC')
    gsc_headers = ['URL', 'Query', 'Impresiones', 'Clics', 'CTR', 'Posición', 'Tipo oportunidad', 'Acción recomendada']
    ws_gsc.append(gsc_headers)
    style_header_row(ws_gsc)

    for i, row in enumerate(gsc_final, start=2):
        ws_gsc.append([row[h] for h in gsc_headers])
        for col_idx in range(1, len(gsc_headers) + 1):
            cell = ws_gsc.cell(row=i, column=col_idx)
            cell.border = thin_border
            col_letter = get_column_letter(col_idx)
            if i % 2 == 0:
                cell.fill = PatternFill('solid', fgColor='F0F4F8')
            if col_letter == 'A':
                cell.font = Font(size=9, name='Calibri', color='1155CC')
                cell.alignment = top_align
            else:
                cell.font = body_font
                cell.alignment = top_align
        ws_gsc.row_dimensions[i].height = 18

    set_col_widths(ws_gsc, {'A': 70, 'B': 35, 'C': 12, 'D': 10, 'E': 10, 'F': 12, 'G': 28, 'H': 50})
    ws_gsc.freeze_panes = 'A2'
    ws_gsc.sheet_properties.tabColor = '27AE60'


    # Guardar
    wb.save(OUTPUT_PATH)
    print(f"\n✅ Excel guardado: {OUTPUT_PATH}")
    print(f"\nResumen final:")
    print(f"  Hoja 'Resumen':           {len(resumen_rows) + 1} filas")
    print(f"  Hoja 'Tareas':            {len(tasks) + 1} filas ({len(tasks)} tareas)")
    print(f"  Hoja 'URLs-Prioridad':    {len(url_prio_final) + 1} filas")
    print(f"  Hoja 'Oportunidades GSC': {len(gsc_final) + 1} filas")

    return {
        'tasks':       len(tasks),
        'urls':        len(url_prio_final),
        'gsc':         len(gsc_final),
        'resumen':     len(resumen_rows),
        'output_path': output_path,
    }


# ──────────────────────────────────────────────────────────────────────────────
# CLI — solo se ejecuta cuando se llama directamente como script
# ──────────────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Motor de auditoría SEO genérico')
    parser.add_argument('--config',     required=True, help='Ruta al archivo config_*.py')
    parser.add_argument('--csv',        help='Ruta al CSV (sobreescribe el del config)')
    parser.add_argument('--output-dir', help='Directorio de salida (sobreescribe el del config)')
    args = parser.parse_args()

    _cfg        = load_config(args.config)
    _ruta_csv   = args.csv or _cfg.RUTA_CSV
    _output_dir = (args.output_dir or _cfg.OUTPUT_DIR).rstrip('/')
    _domain     = _cfg.DOMAIN
    _fecha      = datetime.now().strftime('%Y%m%d')
    _output_path = f"{_output_dir}/auditoria-seo-{_domain}-{_fecha}.xlsx"

    print(f"Config  : {args.config}")
    run_audit(_cfg, _ruta_csv, _output_path)
